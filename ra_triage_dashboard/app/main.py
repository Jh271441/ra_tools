from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import logging
import math
import mimetypes
import re
import shutil
import threading
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

import openpyxl
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import (
    FileResponse,
    HTMLResponse,
    JSONResponse,
    RedirectResponse,
    Response,
)
from fastapi.staticfiles import StaticFiles
from PIL import Image, ImageOps, UnidentifiedImageError

from .assets import AssetIndex, CameraIndex, VideoIndex
from .auth import (
    has_same_origin_mutation_marker,
    identity_header_candidates,
    normalise_username,
    request_identity,
    validate_identity_settings,
)
from .autotriage_source import (
    AutoTriageSource,
    AutoTriageSourceError,
    normalise_batch_id,
)
from .baseline import load_label_baseline
from .batch_prediction_runner import BatchPredictionRunner
from .db import (
    LABELS,
    REVIEW_STATUSES,
    AnnotationConflictError,
    Database,
)
from .model_catalog import MODEL_ID_RE, ModelCatalog, ModelCatalogError
from .prompt_catalog import (
    INPUT_PRESETS,
    MAX_FRAME_COUNT,
    MAX_FRAME_OFFSET_MS,
    MAX_PROMPT_BYTES,
    MIN_FRAME_OFFSET_MS,
    PromptCatalog,
    PromptCatalogError,
    normalise_input_config,
)
from .review_analysis import (
    COMPARISON_STATUSES,
    build_review_reason_analysis,
)
from .work_split import distribute_issue_ids
from .sanitization import redact_sensitive_fields
from .settings import Settings
from .system_status import backup_status, overall_status, volume_status
from .trail_sync import (
    TRAIL_INFO_FIELD,
    TRAIL_RESULT_FIELD,
    read_trail_issue_metadata,
    read_trail_model_fields,
)
from .web_paths import render_index_html, with_base_path


logger = logging.getLogger("ra_triage_dashboard")
_identity_diagnostic_observations: set[tuple[str, tuple[tuple[str, str], ...]]] = set()
settings = Settings.from_env()
validate_identity_settings(settings)
database = Database(
    settings.database_url,
    postgres_migrations_dir=settings.postgres_migrations_dir,
    pool_size=10,
)
asset_index = AssetIndex(
    ra_root=settings.ra_auto_triage_root,
    manifest_path=settings.ares_manifest,
    base_path=settings.base_path,
)
camera_index = CameraIndex(settings.camera_root, base_path=settings.base_path)
video_index = VideoIndex(settings.ares_video_root, base_path=settings.base_path)
model_catalog = ModelCatalog(settings)
prompt_catalog = PromptCatalog(settings.ra_auto_triage_root)
autotriage_source = AutoTriageSource(settings.autotriage_api_base_url)
batch_prediction_runner = BatchPredictionRunner(settings, database)
trail_sync_lock = threading.Lock()
review_image_semaphore = asyncio.Semaphore(2)
thumbnail_image_semaphore = asyncio.Semaphore(4)
trail_detail_semaphore = asyncio.Semaphore(2)
APP_STARTED_AT = datetime.now(timezone.utc)
APP_STARTED_MONOTONIC = time.monotonic()
INDEX_HTML = render_index_html(
    (settings.static_dir / "index.html").read_text(encoding="utf-8"),
    settings.base_path,
)

ISSUE_ID_RE = re.compile(r"^[A-Za-z0-9_-]{3,128}$")
MAX_UPLOAD_BYTES = 64 * 1024 * 1024
MAX_REVIEW_ATTACHMENTS = 4
MAX_REVIEW_ATTACHMENT_BYTES = 8 * 1024 * 1024
MAX_REVIEW_ATTACHMENTS_TOTAL_BYTES = 24 * 1024 * 1024
MAX_REVIEW_MULTIPART_REQUEST_BYTES = 26 * 1024 * 1024
MAX_REVIEW_ATTACHMENT_PIXELS = 40_000_000
MAX_REVIEW_ATTACHMENT_STORAGE_BYTES = 20 * 1024 * 1024 * 1024
MIN_REVIEW_ATTACHMENT_DISK_FREE = 256 * 1024 * 1024
MAX_BATCH_JSON_REQUEST_BYTES = 256 * 1024
MAX_SOURCE_PREVIEW_ROWS = 200
MAX_SOURCE_PREVIEW_CELL_LENGTH = 2_000


def _public_path(path: str) -> str:
    return with_base_path(settings.base_path, path)


MISSING_EVIDENCE_CATALOG: tuple[dict[str, str], ...] = (
    {"key": "routing_direction", "label": "routing 方向缺失", "hint": "未识别自车目标转向 / 车道任务"},
    {"key": "hazard_signal", "label": "双闪缺失", "hint": "未识别前方车辆双闪、临停或故障信号"},
)

REVIEW_TAG_CATALOG: tuple[dict[str, Any], ...] = (
    # Issue description: keep scene context separate from interaction
    # decisions.  The false/true-trigger buckets are retained because they are
    # part of the existing Review vocabulary and historical annotations.
    {"key": "construction_change", "label": "施工/变更区域", "section": "scene", "group": "environment"},
    {"key": "gate", "label": "道闸", "section": "scene", "group": "environment"},
    {"key": "park_entrance", "label": "园区出入口", "section": "scene", "group": "environment"},
    {"key": "environment_u_turn", "label": "掉头", "section": "scene", "group": "environment"},
    {"key": "environment_other", "label": "其他", "section": "scene", "group": "environment"},
    {"key": "intent_straight", "label": "直行", "section": "scene", "group": "self_intent"},
    {"key": "intent_left_turn", "label": "左转", "section": "scene", "group": "self_intent"},
    {"key": "intent_right_turn", "label": "右转", "section": "scene", "group": "self_intent"},
    {"key": "intent_u_turn", "label": "掉头", "section": "scene", "group": "self_intent"},
    {"key": "traffic_light", "label": "等灯", "section": "interaction_decision", "group": "false_trigger"},
    {"key": "queue", "label": "排队", "section": "interaction_decision", "group": "false_trigger"},
    {"key": "yielding", "label": "让行", "section": "interaction_decision", "group": "false_trigger"},
    {"key": "u_turn", "label": "掉头", "section": "interaction_decision", "group": "false_trigger"},
    {"key": "park_in", "label": "泊入", "section": "interaction_decision", "group": "false_trigger"},
    {"key": "park_out", "label": "泊出", "section": "interaction_decision", "group": "false_trigger"},
    {"key": "scene_false_other", "label": "其他", "section": "interaction_decision", "group": "false_trigger"},
    {"key": "obstacle_not_avoided", "label": "未避障", "section": "interaction_decision", "group": "true_trigger"},
    {"key": "close_distance", "label": "距离近", "section": "interaction_decision", "group": "true_trigger"},
    {"key": "perception_fp", "label": "感知FP", "section": "interaction_decision", "group": "true_trigger"},
    {"key": "scene_true_other", "label": "其他", "section": "interaction_decision", "group": "true_trigger"},
    # Issue resolution: how could the vehicle leave the scene?
    {"key": "egress_swag", "label": "SWAG", "section": "egress", "group": "ra"},
    {"key": "egress_detour", "label": "左右绕行", "section": "egress", "group": "ra"},
    {"key": "egress_waypoint", "label": "Waypoint", "section": "egress", "group": "ra"},
    {"key": "egress_reverse", "label": "倒车", "section": "egress", "group": "ra"},
    {"key": "egress_traffic_light", "label": "红绿灯通行", "section": "egress", "group": "ra"},
    {"key": "egress_ra_other", "label": "其他", "section": "egress", "group": "ra"},
    {"key": "lead_vehicle_departed", "label": "前车驶离", "section": "egress", "group": "no_assist"},
    {"key": "system_decision_change", "label": "主系统决策变化", "section": "egress", "group": "no_assist"},
    {"key": "perception_fp_change", "label": "感知FP变化", "section": "egress", "group": "no_assist"},
    {"key": "egress_no_assist_other", "label": "其他", "section": "egress", "group": "no_assist"},
    # Legacy values remain readable in Review history but are no longer offered
    # as new Issue tags.  Keeping them in the contract avoids losing old data.
    {"key": "manual_trigger", "label": "人工触发", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "perception_fp_cleared", "label": "感知FP消失", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "occlusion", "label": "大车遮挡", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "right_turn", "label": "右转", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "left_turn", "label": "左转", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "temporary_stop", "label": "前车双闪", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "vulnerable_road_user", "label": "摩自/行人", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "gt_boundary", "label": "GT 待复核", "section": "legacy", "group": "legacy", "visible": False},
    {"key": "scene_other", "label": "其他（旧交互决策）", "section": "legacy", "group": "legacy", "visible": False},
)
REVIEW_TAG_KEYS = frozenset(item["key"] for item in REVIEW_TAG_CATALOG)
REVIEW_TAG_SCENE_GROUPS = frozenset({"environment", "self_intent"})
REVIEW_TAG_ALIASES = {
    "红绿灯": "traffic_light",
    "等灯": "traffic_light",
    "排队": "queue",
    "让行": "yielding",
    "掉头": "u_turn",
    "施工/变更区域": "construction_change",
    "施工区域": "construction_change",
    "变更区域": "construction_change",
    "道闸": "gate",
    "园区出入口": "park_entrance",
    "场景其他": "environment_other",
    "直行": "intent_straight",
    "自车直行": "intent_straight",
    "自车左转": "intent_left_turn",
    "自车右转": "intent_right_turn",
    "自车掉头": "intent_u_turn",
    "泊入": "park_in",
    "泊出": "park_out",
    "人工触发": "manual_trigger",
    "感知FP消失": "perception_fp_cleared",
    "感知FP": "perception_fp",
    "前车驶离": "lead_vehicle_departed",
    "主系统决策变化": "system_decision_change",
    "未避障": "obstacle_not_avoided",
    "距离近": "close_distance",
    "红绿灯通行": "egress_traffic_light",
    "左右绕行": "egress_detour",
    "Waypoint": "egress_waypoint",
    "倒车": "egress_reverse",
    "感知FP变化": "perception_fp_change",
    "双闪临停": "temporary_stop",
    "前方大车遮挡": "occlusion",
    "大车遮挡": "occlusion",
    "右转": "intent_right_turn",
    "左转": "intent_left_turn",
    "左转待转": "intent_left_turn",
    # Preserve common historical values while the new UI emits the compact catalog above.
    "信号灯": "traffic_light",
    "双闪": "temporary_stop",
    "临停": "temporary_stop",
    "故障车": "temporary_stop",
    "遮挡": "occlusion",
    "摩自": "vulnerable_road_user",
    "行人": "vulnerable_road_user",
    "SWAG": "egress_swag",
    "RA": "egress_swag",
    "GT": "gt_boundary",
    "GT待复核": "gt_boundary",
}

runtime_state: dict[str, Any] = {
    "baseline": {"status": "not_loaded", "message": "等待加载 0508 baseline。", "count": 0},
    "trail_sync": {
        "status": "not_started",
        "message": "尚未检查 Trail 模型字段。",
        "run_id": "",
        "can_create": False,
        "default_changed": False,
    },
}


EXAMPLE_CASES: tuple[dict[str, str], ...] = (
    {
        "issue_id": "cn32171803",
        "title": "左转待转，等灯场景",
        "scenario": "红绿灯周期性等待",
        "summary": "多个路口红灯持续亮起，有停止线；前方摩自停在停止线后方，自车同步等待。",
        "review_note": "当前模型说明为“正确判断为等灯”；用于核验等灯识别与标注流程。",
        "trail_url": "https://voyager.intra.xiaojukeji.com/static/management/#/issue/cn32171803?view_id=2410",
    },
    {
        "issue_id": "cn31954847",
        "title": "排队等灯，前方大车遮挡",
        "scenario": "红绿灯周期性等待",
        "summary": "红灯、停止线/斑马线明确；白色厢式货车停在停止线后，红灯转绿后车流通行。",
        "review_note": "当前模型说明为“正确判断排队等灯”；可用于检验大车遮挡下的等灯识别。",
        "trail_url": "https://voyager.intra.xiaojukeji.com/static/management/#/issue/cn31954847?view_id=2410",
    },
    {
        "issue_id": "cn32000543",
        "title": "自车右转，前方双闪临停车",
        "scenario": "绕行/异常停车",
        "summary": "模型判为排队，未覆盖 RA 协助下绕行通行；案例关注双闪特征。",
        "review_note": "问题假设：双闪缺失导致“排队”FP。请重点标注异常车辆与可绕行性。",
        "trail_url": "https://voyager.intra.xiaojukeji.com/static/management/#/issue/cn32000543?view_id=2410",
    },
    {
        "issue_id": "cn32044177",
        "title": "自车右转，摩自直行且有绕行空间",
        "scenario": "routing 方向 / 绕行空间",
        "summary": "模型判为等灯，但未判断 routing 方向和可绕行空间。",
        "review_note": "问题假设：routing 方向缺失导致“等灯”FP。",
        "trail_url": "https://voyager.intra.xiaojukeji.com/static/management/#/issue/cn32044177?view_id=2410",
    },
    {
        "issue_id": "cn32000563",
        "title": "自车右转，在直行车道排队",
        "scenario": "routing 方向 / 右侧通行空间",
        "summary": "模型判为排队，未识别右侧可右转通行空间；SWAG 右变道后又左加塞回原车道，需复核。",
        "review_note": "问题假设：routing 方向缺失导致“排队”FP；需再 review SWAG 操作链。",
        "trail_url": "https://voyager.intra.xiaojukeji.com/static/management/#/issue/cn32000563?view_id=2410",
    },
    {
        "issue_id": "cn31983487",
        "title": "自车右转，前车直行等灯且无绕行空间",
        "scenario": "routing 方向 / 无绕行空间",
        "summary": "模型判为等灯，但没有判断 routing 方向。",
        "review_note": "问题假设：routing 方向缺失导致“等灯”FP。",
        "trail_url": "https://voyager.intra.xiaojukeji.com/static/management/#/issue/cn31983487?view_id=2410",
    },
)


def _detail(status_code: int, message: str) -> HTTPException:
    return HTTPException(status_code=status_code, detail=message)


def _safe_filename(name: str) -> str:
    result = re.sub(r"[^A-Za-z0-9._-]+", "_", name or "upload")
    return result.strip("._")[:120] or "upload"


def _model_source_artifact_path(source_hash: str, filename: str) -> Path:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".json", ".csv", ".xlsx", ".xlsm"}:
        suffix = ".bin"
    return settings.uploads_dir / f"model-run-{source_hash}{suffix}"


def _store_model_source(content: bytes, filename: str, source_hash: str) -> dict[str, str]:
    root = settings.uploads_dir.resolve()
    root.mkdir(parents=True, exist_ok=True)
    path = _model_source_artifact_path(source_hash, filename).resolve()
    if root not in path.parents:
        raise OSError("模型结果来源文件路径越界。")
    if not path.is_file():
        path.write_bytes(content)
    media_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    return {
        "stored_name": path.name,
        "filename": _safe_filename(filename),
        "media_type": media_type,
    }


def _model_source_file(run: dict[str, Any]) -> tuple[Path, str] | None:
    metadata = run.get("metadata") if isinstance(run.get("metadata"), dict) else {}
    artifact = metadata.get("source_artifact") if isinstance(metadata.get("source_artifact"), dict) else {}
    stored_name = _as_text(artifact.get("stored_name"))
    if stored_name:
        root = settings.uploads_dir.resolve()
        candidate = (root / stored_name).resolve()
        if root in candidate.parents and candidate.is_file():
            return candidate, _safe_filename(_as_text(artifact.get("filename")) or candidate.name)

    # Runs created before source_artifact metadata was introduced can still be
    # resolved by their immutable content hash when the archive was written.
    source_hash = _as_text(run.get("source_sha256"))
    source_name = _as_text(run.get("source_name"))
    if source_hash and run.get("kind", "upload") == "upload":
        root = settings.uploads_dir.resolve()
        candidate = _model_source_artifact_path(source_hash, source_name).resolve()
        if root in candidate.parents and candidate.is_file():
            return candidate, _safe_filename(Path(source_name).name or candidate.name)

    # Older runs may point at a source file that was already present on the
    # server. Keep this fallback constrained to known dashboard/data roots.
    if source_name:
        candidate = Path(source_name).expanduser()
        allowed_roots = (
            settings.uploads_dir.resolve(),
            settings.data_dir.resolve(),
            settings.ra_auto_triage_root.resolve(),
        )
        if candidate.is_absolute() and candidate.is_file():
            resolved = candidate.resolve()
            if any(root == resolved or root in resolved.parents for root in allowed_roots):
                return resolved, _safe_filename(resolved.name)
    return None


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def _source_preview_value(value: Any) -> str:
    """Return a bounded, credential-redacted cell value for the preview UI."""

    safe_value = redact_sensitive_fields(value)
    if isinstance(safe_value, (dict, list, tuple)):
        text = json.dumps(
            safe_value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            default=str,
        )
    else:
        text = _as_text(safe_value)
    if len(text) <= MAX_SOURCE_PREVIEW_CELL_LENGTH:
        return text
    return text[: MAX_SOURCE_PREVIEW_CELL_LENGTH - 1] + "…"


def _model_source_filename(run: dict[str, Any]) -> str:
    metadata = run.get("metadata") if isinstance(run.get("metadata"), dict) else {}
    artifact = metadata.get("source_artifact") if isinstance(metadata.get("source_artifact"), dict) else {}
    return _safe_filename(
        _as_text(artifact.get("filename"))
        or Path(_as_text(run.get("source_name"))).name
        or "model-results.json"
    )


def _reconstructed_model_source(
    run_id: str,
    run: dict[str, Any],
) -> tuple[bytes, str, str] | None:
    """Rebuild a safe source copy for uploads made before file archiving."""

    filename = _model_source_filename(run)
    suffix = Path(filename).suffix.lower()
    if suffix not in {".json", ".csv"}:
        return None
    rows = database.model_run_source_rows(run_id)
    if not rows:
        return None
    if suffix == ".csv":
        columns: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    columns.append(key)
        output = io.StringIO(newline="")
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig"), filename, "text/csv; charset=utf-8"
    payload = {
        "schema_version": "reconstructed-model-run-v1",
        "source_name": _as_text(run.get("source_name")),
        "source_sha256": _as_text(run.get("source_sha256")),
        "results": rows,
    }
    return (
        json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode("utf-8"),
        filename,
        "application/json",
    )


def _voyager_issue_url(issue_id: str) -> str:
    return (
        f"{settings.voyager_issue_base_url.rstrip('/')}/"
        f"{quote(issue_id, safe='')}?view_id={settings.voyager_issue_view_id}"
    )


def _ra_recording_url(ra_id: Any) -> str:
    """Build the read-only RA dashboard URL from Trail's canonical ra_id."""

    value = _as_text(ra_id)
    if not value:
        return ""
    return (
        f"{settings.ra_recording_base_url.rstrip('/')}/"
        f"{quote(value, safe='')}?returnUrl="
    )


def _case_external_links(issue_id: str, metadata: dict[str, Any]) -> dict[str, Any]:
    """Expose the small, read-only RA link/event subset to the browser."""

    events = metadata.get("ra_event")
    if isinstance(events, str):
        try:
            events = json.loads(events)
        except (TypeError, ValueError, json.JSONDecodeError):
            events = []
    safe_events: list[dict[str, Any]] = []
    if isinstance(events, (list, tuple)):
        for item in events:
            if not isinstance(item, dict):
                continue
            event = str(item.get("event") or "").strip()[:128]
            if not event:
                continue
            raw_timestamp = item.get("timestamp")
            try:
                timestamp = int(raw_timestamp) if raw_timestamp is not None else None
            except (TypeError, ValueError, OverflowError):
                timestamp = None
            raw_value = item.get("value")
            if isinstance(raw_value, (str, int, float, bool)) or raw_value is None:
                value = raw_value
            else:
                value = str(raw_value)[:256]
            safe_events.append({"event": event, "value": value, "timestamp": timestamp})
    safe_events = safe_events[:64]
    event_count = len(safe_events)
    return {
        "ra_recording_url": _ra_recording_url(metadata.get("ra_id")),
        "ra_event_url": _voyager_issue_url(issue_id) if event_count else "",
        "ra_task_id": _as_text(metadata.get("ra_id")),
        "ra_event_count": event_count,
        "ra_events": safe_events,
    }


def _case_link_metadata_fallback(case: dict[str, Any]) -> dict[str, Any]:
    """Use explicitly imported RA fields when Trail detail lookup is disabled."""

    fallback: dict[str, Any] = {}
    for source in (case, case.get("extra")):
        if not isinstance(source, dict):
            continue
        for key in ("ra_id", "ra_event"):
            if key not in fallback and source.get(key) not in (None, "", []):
                fallback[key] = source[key]
    return fallback


def _autotriage_record_url(batch_id: str) -> str:
    if not batch_id:
        return ""
    return (
        f"{settings.auto_triage_record_base_url.rstrip('/')}/"
        f"{quote(batch_id, safe='')}?tab=results"
    )


def _public_batch_job(
    job: dict[str, Any],
    *,
    include_prompt: bool = False,
) -> dict[str, Any]:
    public = dict(job)
    prompt_template = _as_text(public.get("prompt_template"))
    public["prompt_template_length"] = len(prompt_template)
    public["prompt_preview"] = re.sub(r"\s+", " ", prompt_template)[:240]
    if not include_prompt:
        public.pop("prompt_template", None)
    public["record_url"] = _autotriage_record_url(
        _as_text(job.get("autotriage_batch_id"))
    )
    if "items" in job:
        public["items"] = [
            {
                **item,
                "voyager_issue_url": _voyager_issue_url(
                    _as_text(item.get("issue_id"))
                ),
            }
            for item in job.get("items", [])
        ]
    return public


def _safe_autotriage_batch(batch: dict[str, Any]) -> dict[str, Any]:
    allowed = {
        "id",
        "batch_name",
        "username",
        "status",
        "prompt_version",
        "prompt_text",
        "model_name",
        "total_count",
        "completed_count",
        "success_count",
        "failed_count",
        "match_count",
        "mismatch_count",
        "accuracy",
        "experiment_id",
        "experiment_revision_id",
        "experiment_source",
        "created_at",
        "started_at",
        "finished_at",
    }
    safe = redact_sensitive_fields(
        {key: batch.get(key) for key in allowed if key in batch}
    )
    prompt_text = _as_text(safe.pop("prompt_text", ""))
    safe["prompt_sha256"] = (
        hashlib.sha256(prompt_text.encode("utf-8")).hexdigest()
        if prompt_text
        else ""
    )
    safe["prompt_preview"] = re.sub(r"\s+", " ", prompt_text)[:240]
    safe["prompt_length"] = len(prompt_text)
    return safe


def _thumbnail_cache_path(issue_id: str, source: Path) -> Path:
    stat = source.stat()
    fingerprint = (
        f"{issue_id}\0{source}\0{stat.st_mtime_ns}\0{stat.st_size}"
    ).encode("utf-8")
    digest = hashlib.sha256(fingerprint).hexdigest()
    return settings.case_thumbnails_dir / f"{digest}.jpg"


def _render_case_thumbnail(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(source) as opened:
        image = ImageOps.exif_transpose(opened)
        if image.width * image.height > MAX_REVIEW_ATTACHMENT_PIXELS:
            raise ValueError("BEV 图片像素数过大。")
        if image.mode != "RGB":
            image = image.convert("RGB")
        contained = ImageOps.contain(
            image,
            (640, 360),
            method=Image.Resampling.LANCZOS,
        )
        canvas = Image.new("RGB", (640, 360), color=(11, 18, 32))
        canvas.paste(
            contained,
            ((640 - contained.width) // 2, (360 - contained.height) // 2),
        )
        temp_path = destination.with_name(
            f".{destination.name}.{uuid.uuid4().hex}.tmp"
        )
        try:
            canvas.save(
                temp_path,
                format="JPEG",
                quality=78,
                optimize=True,
                progressive=True,
            )
            temp_path.replace(destination)
        finally:
            temp_path.unlink(missing_ok=True)


def _action_actor(request: Request, submitted_name: Any = "") -> tuple[str, str, bool]:
    """Resolve an audit/display actor without trusting direct-IP client claims."""

    identity = request_identity(request, settings)
    if identity.verified and identity.username:
        return identity.username, identity.source, True
    username = normalise_username(_as_text(submitted_name))
    if username:
        return username, "client_claim_unverified", False
    return "", "anonymous", False


def _can_manage_team_default(request: Request) -> bool:
    identity = request_identity(request, settings)
    return bool(
        identity.verified
        and identity.username
        and database.access_role(identity.username) == "admin"
    )


def _admin_identity(request: Request):
    identity = request_identity(request, settings)
    if not (
        identity.verified
        and identity.username
        and database.access_role(identity.username) == "admin"
    ):
        raise _detail(403, "该操作仅限 Dashboard 管理员。")
    return identity


def _review_tag_catalog() -> tuple[dict[str, Any], ...]:
    """Return built-in tags plus shared scene tags from the database.

    Built-ins remain source-controlled, but a database row can override their
    label/hint/group or soft-delete them (same pattern as missing evidence).
    """

    merged: dict[str, dict[str, Any]] = {
        str(item["key"]): {
            **item,
            "builtin": True,
            "deleted": False,
        }
        for item in REVIEW_TAG_CATALOG
    }
    builtin_keys = set(merged)
    for row in database.list_review_tag_catalog(include_inactive=True):
        key = str(row.get("key") or "").strip()
        if not key:
            continue
        item = {
            str(name): value
            for name, value in row.items()
            if str(name) != "active"
        }
        if "group_key" in item and "group" not in item:
            item["group"] = item.pop("group_key")
        item.setdefault("section", "scene")
        item.setdefault("group", "environment")
        item.setdefault("hint", "")
        item["builtin"] = key in builtin_keys
        item["deleted"] = not bool(row.get("active", 1))
        if key in merged:
            merged[key].update(item)
            merged[key]["builtin"] = True
        else:
            merged[key] = item
    return tuple(merged.values())


def _missing_evidence_catalog() -> tuple[dict[str, Any], ...]:
    """Return the merged catalog, including soft-deleted historical entries.

    Built-ins remain source-controlled, but a database row can override their
    label/hint or retire them.  Retired entries stay in the payload so old
    annotations remain readable; the Review form hides them unless selected
    by the current version.
    """

    merged: dict[str, dict[str, Any]] = {
        str(item["key"]): {
            **item,
            "builtin": True,
            "deleted": False,
        }
        for item in MISSING_EVIDENCE_CATALOG
    }
    builtin_keys = set(merged)
    for row in database.list_missing_evidence_catalog(include_inactive=True):
        key = str(row.get("key") or "").strip()
        if not key:
            continue
        item = {
            str(name): value
            for name, value in row.items()
            if str(name) != "active"
        }
        item["builtin"] = key in builtin_keys
        item["deleted"] = not bool(row.get("active", 1))
        if key in merged:
            merged[key].update(item)
            merged[key]["builtin"] = True
        else:
            merged[key] = item
    return tuple(merged.values())


def _value(row: dict[str, Any], *names: str) -> Any:
    lower = {str(key).strip().lower(): value for key, value in row.items()}
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
        value = lower.get(name.lower())
        if value not in (None, ""):
            return value
    return ""


def _parse_structured(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if isinstance(value, list):
        return {"items": value}
    if not isinstance(value, str) or not value.strip():
        return {}
    try:
        parsed = json.loads(value)
        if isinstance(parsed, dict):
            return parsed
        if isinstance(parsed, list):
            return {"items": parsed}
    except (TypeError, ValueError):
        pass
    return {"text": value.strip()}


def _number_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _canonical_gt_label(value: Any) -> str:
    text = _as_text(value)
    mapping = {
        "false_positive": "误触发",
        "fp": "误触发",
        "false positive": "误触发",
        "true_positive": "正确触发",
        "tp": "正确触发",
        "true positive": "正确触发",
        "no_assist": "无需协助",
        "no_assistance": "无需协助",
        "不需要协助": "无需协助",
        "不需协助": "无需协助",
        "无需远程协助": "无需协助",
        "无需远程辅助": "无需协助",
        "无需人工协助": "无需协助",
    }
    text = mapping.get(text.lower(), text)
    return text if text in LABELS else ""


def _label_from_structured(value: dict[str, Any]) -> str:
    for key in ("label", "model_label", "result", "prediction", "triage_result", "class"):
        candidate = _canonical_gt_label(value.get(key))
        if candidate:
            return candidate
    return ""


def _first_text(value: dict[str, Any], *keys: str) -> str:
    for key in keys:
        text = _as_text(value.get(key))
        if text:
            return text
    return ""


def normalize_model_row(row: dict[str, Any]) -> dict[str, Any] | None:
    issue_id = _as_text(
        _value(row, "issue_id", "issueId", "issue", "问题id", "问题ID")
    )
    if not ISSUE_ID_RE.fullmatch(issue_id):
        return None
    raw_result = _value(
        row,
        "model_label",
        "ra_stuck_auto_result",
        "prediction",
        "pred_label",
        "预测标签",
    )
    result_info = _parse_structured(raw_result)
    info = _parse_structured(_value(row, "ra_stuck_auto_result_info", "result_info"))
    raw_label = _as_text(raw_result)
    label = _canonical_gt_label(raw_label) or _label_from_structured(result_info) or _label_from_structured(info)
    # Preserve a nonstandard class for audit/debugging.  It will be visibly
    # marked as non-comparable rather than being silently coerced to a GT label.
    if not label and raw_label and not isinstance(raw_result, (dict, list)):
        label = raw_label
    reason = _as_text(
        _value(
            row,
            "model_reason",
            "reason",
            "预测理由",
            "ra_stuck_auto_reason",
        )
    ) or _first_text(info, "reason", "model_reason", "analysis", "explanation", "text")
    confidence = _number_or_none(
        _value(row, "model_confidence", "confidence", "置信度")
        or info.get("confidence")
        or info.get("model_confidence")
    )
    extra = _value(row, "model_extra")
    if not isinstance(extra, dict):
        extra = {}
    if info:
        extra = {**extra, "ra_stuck_auto_result_info": info}
    if result_info and result_info != info:
        extra["ra_stuck_auto_result_payload"] = result_info
    return {
        "issue_id": issue_id,
        "trip_id": _as_text(_value(row, "trip_id", "tripId")),
        "model_label": label,
        "model_reason": reason,
        "model_confidence": confidence,
        "model_extra": extra,
        "raw": row,
    }


def _fetch_autotriage_snapshot(batch_ref: Any) -> dict[str, Any]:
    batch_id = normalise_batch_id(batch_ref)
    batch = autotriage_source.fetch_batch(batch_id)
    source_rows = autotriage_source.fetch_results(batch_id)
    safe_batch = _safe_autotriage_batch(batch)
    rows: list[dict[str, Any]] = []
    rejected = 0
    seen: set[str] = set()
    for source_row in source_rows:
        redacted_row = redact_sensitive_fields(source_row)
        normalized = normalize_model_row(redacted_row)
        explicit_success = source_row.get("success")
        row_failed = explicit_success is False or (
            isinstance(explicit_success, str)
            and explicit_success.strip().lower() in {"false", "0", "failed"}
        )
        if (
            row_failed
            or normalized is None
            or normalized.get("model_label") not in LABELS
        ):
            rejected += 1
            continue
        issue_id = _as_text(normalized.get("issue_id"))
        if issue_id in seen:
            raise AutoTriageSourceError(
                f"AutoTriage Batch {batch_id} 含重复 Issue：{issue_id}。"
            )
        seen.add(issue_id)
        normalized["raw"] = redacted_row
        rows.append(normalized)
    if not rows:
        raise AutoTriageSourceError(
            "该 AutoTriage Batch 没有可导入的三分类预测结果。"
        )

    def platform_count(field: str) -> int:
        try:
            count = int(batch.get(field) or 0)
        except (TypeError, ValueError):
            raise AutoTriageSourceError(
                f"AutoTriage Batch 的 {field} 非法。"
            )
        if count < 0:
            raise AutoTriageSourceError(
                f"AutoTriage Batch 的 {field} 不能为负数。"
            )
        return count

    declared_total = platform_count("total_count")
    completed_total = platform_count("completed_count")
    failed_total = platform_count("failed_count")
    platform_status = _as_text(batch.get("status")).lower()
    partial = bool(
        rejected
        or failed_total
        or platform_status not in {"completed", "succeeded"}
        or (declared_total and len(source_rows) != declared_total)
        or (declared_total and completed_total != declared_total)
        or (declared_total and len(rows) != declared_total)
    )
    fingerprint_rows = [
        {
            "issue_id": row["issue_id"],
            "trip_id": row["trip_id"],
            "model_label": row["model_label"],
            "model_reason": row["model_reason"],
            "model_confidence": row["model_confidence"],
            "model_extra": row["model_extra"],
        }
        for row in sorted(rows, key=lambda item: item["issue_id"])
    ]
    fingerprint = {
        "schema_version": "autotriage-snapshot-v1",
        "batch_id": batch_id,
        "batch": safe_batch,
        "predictions": fingerprint_rows,
        "source_result_count": len(source_rows),
        "rejected_result_count": rejected,
    }
    source_sha256 = hashlib.sha256(
        json.dumps(
            fingerprint,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
    ).hexdigest()
    coverage = {
        "declared_total": declared_total,
        "completed_total": completed_total,
        "failed_total": failed_total,
        "platform_status": platform_status,
        "source_result_count": len(source_rows),
        "accepted_result_count": len(rows),
        "rejected_result_count": rejected,
        "unique_issue_count": len(seen),
        "partial": partial,
    }
    return {
        "batch_id": batch_id,
        "batch": safe_batch,
        "rows": rows,
        "coverage": coverage,
        "source_sha256": source_sha256,
    }


def normalize_issue_row(row: dict[str, Any], source: str) -> dict[str, Any] | None:
    issue_id = _as_text(
        _value(row, "issue_id", "issueId", "issue", "问题id", "问题ID")
    )
    if not ISSUE_ID_RE.fullmatch(issue_id):
        return None
    gt_label = _canonical_gt_label(
        _value(
            row,
            "gt_label",
            "ra_merge_result",
            "期望输出",
            "真实标签",
            "ground_truth",
        )
    )
    return {
        "issue_id": issue_id,
        "trip_id": _as_text(_value(row, "trip_id", "tripId")),
        "title": _as_text(_value(row, "title", "标题", "名称")),
        "scenario": _as_text(_value(row, "scenario", "场景")),
        "summary": _as_text(_value(row, "summary", "描述", "description")),
        "review_note": _as_text(_value(row, "review_note", "备注", "note")),
        "trail_url": _as_text(_value(row, "trail_url", "url", "链接")),
        "gt_label": gt_label,
        "gt_source": source if gt_label else "",
        "extra": redact_sensitive_fields(row),
    }


def parse_source_bytes(filename: str, content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".json":
        parsed = json.loads(content.decode("utf-8-sig"))
        if isinstance(parsed, dict):
            rows = parsed.get("results") or parsed.get("data") or parsed.get("rows") or []
            metadata = {
                key: value
                for key, value in parsed.items()
                if key not in {"results", "data", "rows"}
            }
        elif isinstance(parsed, list):
            rows, metadata = parsed, {}
        else:
            raise ValueError("JSON 顶层必须是对象或数组。")
        if not isinstance(rows, list):
            raise ValueError("JSON 的 results/data/rows 必须是数组。")
        return [
            row for row in rows if isinstance(row, dict)
        ], redact_sensitive_fields(metadata)

    if suffix == ".csv":
        text = ""
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if not text:
            raise ValueError("CSV 编码无法识别，请使用 UTF-8 或 GB18030。")
        return list(csv.DictReader(io.StringIO(text))), {}

    if suffix in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            raise ValueError("Excel 缺少表头。")
        columns = [str(value).strip() if value is not None else "" for value in headers]
        rows: list[dict[str, Any]] = []
        for values_row in values:
            if not any(value is not None and str(value).strip() for value in values_row):
                continue
            rows.append(
                {
                    columns[index]: value
                    for index, value in enumerate(values_row)
                    if index < len(columns) and columns[index]
                }
            )
        return rows, {"sheet": sheet.title}

    raise ValueError("仅支持 .json、.csv、.xlsx 或 .xlsm。")


def _normalise_review_image(content: bytes) -> tuple[bytes, str, str, int, int]:
    try:
        with Image.open(io.BytesIO(content)) as source:
            source.seek(0)
            width, height = source.size
            if width <= 0 or height <= 0 or width * height > MAX_REVIEW_ATTACHMENT_PIXELS:
                raise _detail(400, "截图尺寸非法或像素数超过 4000 万。")
            image = ImageOps.exif_transpose(source)
            width, height = image.size
            image_format = (source.format or "").upper()
            output = io.BytesIO()
            if image_format == "PNG":
                if image.mode not in {"1", "L", "LA", "P", "RGB", "RGBA"}:
                    image = image.convert("RGBA")
                image.save(output, format="PNG", optimize=True)
                media_type, suffix = "image/png", ".png"
            elif image_format in {"JPEG", "JPG"}:
                if image.mode != "RGB":
                    image = image.convert("RGB")
                image.save(output, format="JPEG", quality=92, optimize=True)
                media_type, suffix = "image/jpeg", ".jpg"
            elif image_format == "WEBP":
                if image.mode not in {"RGB", "RGBA"}:
                    image = image.convert("RGBA")
                image.save(output, format="WEBP", quality=92, method=4)
                media_type, suffix = "image/webp", ".webp"
            else:
                raise _detail(400, "截图仅支持 PNG、JPEG 或 WebP。")
            normalized = output.getvalue()
    except HTTPException:
        raise
    except (Image.DecompressionBombError, UnidentifiedImageError, OSError, ValueError):
        raise _detail(400, "截图无法解码或文件已损坏。")
    if not normalized or len(normalized) > MAX_REVIEW_ATTACHMENT_BYTES:
        raise _detail(413, "规范化后的单张截图不能超过 8 MB。")
    return normalized, media_type, suffix, width, height


async def _store_review_attachments(
    uploads: list[UploadFile],
) -> tuple[list[dict[str, Any]], list[Path]]:
    if len(uploads) > MAX_REVIEW_ATTACHMENTS:
        raise _detail(400, f"每次最多粘贴 {MAX_REVIEW_ATTACHMENTS} 张截图。")
    prepared: list[tuple[dict[str, Any], bytes]] = []
    raw_total_bytes = 0
    total_bytes = 0
    for upload in uploads:
        content = await upload.read(MAX_REVIEW_ATTACHMENT_BYTES + 1)
        if not content:
            raise _detail(400, "截图文件为空。")
        if len(content) > MAX_REVIEW_ATTACHMENT_BYTES:
            raise _detail(413, "单张截图不能超过 8 MB。")
        raw_total_bytes += len(content)
        if raw_total_bytes > MAX_REVIEW_ATTACHMENTS_TOTAL_BYTES:
            raise _detail(413, "本次截图总大小不能超过 24 MB。")
        async with review_image_semaphore:
            normalized, media_type, suffix, width, height = await asyncio.to_thread(
                _normalise_review_image,
                content,
            )
        total_bytes += len(normalized)
        if total_bytes > MAX_REVIEW_ATTACHMENTS_TOTAL_BYTES:
            raise _detail(413, "本次截图总大小不能超过 24 MB。")
        attachment_id = str(uuid.uuid4())
        stored_name = f"{attachment_id[:2]}/{attachment_id}{suffix}"
        prepared.append(
            (
                {
                    "id": attachment_id,
                    "original_name": _safe_filename(
                        upload.filename or f"clipboard{suffix}"
                    ),
                    "stored_name": stored_name,
                    "media_type": media_type,
                    "size_bytes": len(normalized),
                    "width": width,
                    "height": height,
                    "sha256": hashlib.sha256(normalized).hexdigest(),
                },
                normalized,
            )
        )

    root = settings.review_attachments_dir.resolve()
    root.mkdir(parents=True, exist_ok=True)
    if (
        prepared
        and database.review_attachment_storage_bytes() + total_bytes
        > MAX_REVIEW_ATTACHMENT_STORAGE_BYTES
    ):
        raise _detail(507, "Review 截图已达到 20 GB 存储配额，请联系管理员清理或扩容。")
    if prepared and shutil.disk_usage(root).free < total_bytes + MIN_REVIEW_ATTACHMENT_DISK_FREE:
        raise _detail(507, "截图存储空间不足，请联系管理员。")
    temp_paths: list[Path] = []
    final_paths: list[Path] = []
    try:
        for record, content in prepared:
            stored_name = str(record["stored_name"])
            path = (settings.review_attachments_dir / stored_name).resolve()
            if root not in path.parents:
                raise _detail(400, "截图存储路径非法。")
            path.parent.mkdir(parents=True, exist_ok=True)
            temp_path = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
            temp_path.write_bytes(content)
            temp_paths.append(temp_path)
            final_paths.append(path)
        for temp_path, final_path in zip(temp_paths, final_paths):
            temp_path.replace(final_path)
    except Exception:
        for path in [*temp_paths, *final_paths]:
            path.unlink(missing_ok=True)
        raise
    return [record for record, _ in prepared], final_paths


def _public_review_attachment(attachment: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": attachment["id"],
        "media_type": attachment["media_type"],
        "size_bytes": int(attachment["size_bytes"]),
        "width": int(attachment["width"]),
        "height": int(attachment["height"]),
        "created_at": attachment.get("created_at"),
        "url": _public_path(f"/api/review-attachments/{attachment['id']}"),
    }


def _normalise_review_tags(values: list[Any]) -> list[str]:
    if len(values) > 24:
        raise _detail(400, "每条 review 最多选择 24 个 tags。")
    catalog_keys = {str(item["key"]) for item in _review_tag_catalog()}
    normalized: set[str] = set()
    for value in values:
        raw = str(value).strip()
        if not raw:
            continue
        if len(raw) > 48 or re.search(r"[\x00-\x1f\x7f]", raw):
            raise _detail(400, "tag 长度或字符不合法。")
        key = REVIEW_TAG_ALIASES.get(raw, raw)
        if key in catalog_keys:
            normalized.add(key)
        else:
            raise _detail(400, "tag 不在默认场景 Tags 目录中。")
    return [item["key"] for item in _review_tag_catalog() if item["key"] in normalized]


def _normalise_missing_evidence(values: list[Any]) -> list[str]:
    if len(values) > 24:
        raise _detail(400, "每条 review 最多选择 24 个缺失信息。")
    catalog_keys = {str(item["key"]) for item in _missing_evidence_catalog()}
    normalized: list[str] = []
    seen: set[str] = set()
    for value in values:
        raw = str(value).strip()
        if not raw or raw in seen:
            continue
        if len(raw) > 160 or re.search(r"[\x00-\x1f\x7f]", raw):
            raise _detail(400, "缺失信息字段长度或字符不合法。")
        # Legacy per-Review custom values remain readable and editable. New
        # values are opaque keys from the shared catalog above.
        if raw not in catalog_keys and not raw.startswith("custom:"):
            raise _detail(400, "缺失信息不在共享目录中。")
        seen.add(raw)
        normalized.append(raw)
    return normalized


def _normalise_review_excluded(value: Any) -> bool:
    """Parse the explicit Issue-level exclusion flag without truthiness traps."""

    if value is None or value is False or value == 0:
        return False
    if value is True or value == 1:
        return True
    if isinstance(value, str):
        normalized = value.strip().casefold()
        if normalized in {"", "0", "false", "no", "否", "不排除"}:
            return False
        if normalized in {"1", "true", "yes", "是", "排除"}:
            return True
    raise _detail(400, "is_excluded 必须是布尔值。")


def _create_annotation_record(
    *,
    issue_id: str,
    request: Request,
    body: dict[str, Any],
    attachments: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    label = _as_text(body.get("label"))
    review_status = _as_text(body.get("review_status") or "pending")
    is_excluded = _normalise_review_excluded(body.get("is_excluded", False))
    tags = body.get("tags") or []
    missing_evidence = body.get("missing_evidence") or []
    if not isinstance(tags, list):
        raise _detail(400, "tags 必须是数组。")
    if not isinstance(missing_evidence, list):
        raise _detail(400, "missing_evidence 必须是数组。")
    tags = _normalise_review_tags(tags)
    missing_evidence = _normalise_missing_evidence(missing_evidence)
    model_run_id = _as_text(body.get("model_run_id"))
    has_expected_previous = "expected_previous_annotation_id" in body
    expected_previous_annotation_id: int | None = None
    if has_expected_previous:
        raw_expected = body.get("expected_previous_annotation_id")
        if raw_expected in (None, "", 0, "0"):
            expected_previous_annotation_id = None
        else:
            try:
                expected_previous_annotation_id = int(raw_expected)
            except (TypeError, ValueError) as exc:
                raise _detail(400, "expected_previous_annotation_id 不合法。") from exc
            if expected_previous_annotation_id <= 0:
                raise _detail(400, "expected_previous_annotation_id 不合法。")
    author, author_source, author_verified = _action_actor(request, body.get("author"))
    annotation_kwargs: dict[str, Any] = {
        "issue_id": issue_id,
        "model_run_id": model_run_id,
        "label": label,
        "review_status": review_status,
        "is_excluded": is_excluded,
        "tags": tags,
        "missing_evidence": missing_evidence,
        "note": _as_text(body.get("note")),
        "author": author,
        "author_source": author_source,
        "author_verified": author_verified,
        "attachments": attachments,
    }
    if has_expected_previous:
        annotation_kwargs["expected_previous_annotation_id"] = expected_previous_annotation_id
    try:
        return database.create_annotation(
            **annotation_kwargs,
        )
    except AnnotationConflictError as exc:
        raise _detail(409, str(exc))
    except ValueError as exc:
        raise _detail(400, str(exc))


def bootstrap_model_result() -> None:
    path = settings.bootstrap_model_json
    if not path or not path.is_file():
        return
    content = path.read_bytes()
    try:
        source_rows, metadata = parse_source_bytes(path.name, content)
        rows = [normalized for row in source_rows if (normalized := normalize_model_row(row))]
        if not rows:
            return
        experiment = metadata.get("experiment") if isinstance(metadata.get("experiment"), dict) else {}
        name = _as_text(experiment.get("model_name")) or path.stem
        database.import_model_run(
            name=name,
            source_name=str(path),
            source_sha256=hashlib.sha256(content).hexdigest(),
            metadata=metadata,
            rows=rows,
            created_by="system",
            created_by_source="service_bootstrap",
            created_by_verified=False,
        )
    except Exception as exc:
        print(f"[ra_triage_dashboard] bootstrap model result skipped: {exc}", flush=True)


def bootstrap_baseline() -> dict[str, Any]:
    loaded = load_label_baseline(settings.baseline_label_xlsx, settings.baseline_dataset)
    state = {
        "scope": settings.baseline_scope,
        "dataset": settings.baseline_dataset,
        "path": str(settings.baseline_label_xlsx),
        "source_rows": loaded.source_rows,
        "count": len(loaded.rows),
        "skipped_rows": loaded.skipped_rows,
        "message": loaded.message,
        "status": "ready" if loaded.rows else "unavailable",
    }
    if loaded.rows:
        state["upsert"] = database.replace_baseline_scope(
            scope=settings.baseline_scope,
            rows=loaded.rows,
            source="trail_label_baseline",
        )
    runtime_state["baseline"] = state
    return state


def sync_trail_model_fields(
    *,
    create_run: bool = False,
    requested_by: str = "",
    identity_source: str = "anonymous",
    identity_verified: bool = False,
    trigger: str = "manual",
) -> dict[str, Any]:
    """Inspect Trail fields and optionally create/reuse an immutable local snapshot.

    This function only calls the Trail query client.  Creating a snapshot writes
    local SQLite rows, but never writes Trail and never changes the shared
    default model run.
    """

    if not trail_sync_lock.acquire(blocking=False):
        return {
            **runtime_state["trail_sync"],
            "status": "running",
            "message": "已有 Trail 字段检查或快照任务在运行。",
        }
    action = "创建只读快照" if create_run else "检查字段"
    runtime_state["trail_sync"] = {
        "status": "running",
        "message": f"正在从 Trail view {settings.trail_view_id} 只读{action}。",
        "run_id": "",
        "can_create": False,
        "default_changed": False,
        "action": "create" if create_run else "preview",
    }
    try:
        issue_ids = database.baseline_issue_ids(settings.baseline_scope)
        result = read_trail_model_fields(
            ra_root=settings.ra_auto_triage_root,
            issue_ids=issue_ids,
            view_id=settings.trail_view_id,
            chunk_size=settings.trail_sync_chunk_size,
        )
        state: dict[str, Any] = {
            "status": "unavailable",
            "message": result.message,
            "view_id": result.view_id,
            "queried_issues": result.queried_issues,
            "returned_issues": result.returned_issues,
            "fields_visible": list(result.fields_visible),
            "run_id": "",
            "complete": result.complete,
            "can_create": False,
            "default_changed": False,
            "action": "create" if create_run else "preview",
        }
        if not result.complete or result.returned_issues < result.queried_issues:
            state.update(
                {
                    "status": "failed",
                    "message": (
                        result.message
                        + (
                            f" 仅返回 {result.returned_issues} / {result.queried_issues} 个 baseline issue；"
                            if result.complete
                            else " "
                        )
                        + "为避免部分快照，未创建 Run，团队默认 Run 未变化。"
                    ),
                }
            )
            runtime_state["trail_sync"] = state
            return state
        if TRAIL_RESULT_FIELD not in result.fields_visible:
            state["message"] = result.message + " 未创建 Run，团队默认 Run 未变化。"
            runtime_state["trail_sync"] = state
            return state
        normalized = [row for raw in result.rows if (row := normalize_model_row(raw))]
        # Trail snapshots participate in three-class evaluation, so only the
        # canonical labels are usable. Keep non-standard values out of the
        # snapshot rather than treating placeholders such as "nan" as labels.
        usable = [row for row in normalized if row["model_label"] in LABELS]
        if not usable:
            state.update(
                {
                    "status": "empty",
                    "message": (
                        result.message
                        + " 字段已出现，但当前 baseline 没有非空模型结果；未创建 Run。"
                    ),
                }
            )
            runtime_state["trail_sync"] = state
            return state
        snapshot_rows = sorted(usable, key=lambda row: row["issue_id"])
        state.update(
            {
                "can_create": True,
                "usable_predictions": len(usable),
                "missing_predictions": max(result.queried_issues - len(usable), 0),
            }
        )
        if not create_run:
            state.update(
                {
                    "status": "preview_ready",
                    "message": (
                        result.message
                        + f" 检查完成：{len(usable)} / {result.queried_issues} 条可生成快照。"
                        + " 尚未创建 Run，团队默认 Run 未变化。"
                    ),
                }
            )
            runtime_state["trail_sync"] = state
            return state
        payload = {
            "scope": settings.baseline_scope,
            "view_id": settings.trail_view_id,
            "fields": [TRAIL_RESULT_FIELD, TRAIL_INFO_FIELD],
            "rows": snapshot_rows,
        }
        source_hash = hashlib.sha256(
            json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()
        run, duplicate = database.import_model_run(
            name=f"Trail view {settings.trail_view_id} · {TRAIL_RESULT_FIELD}",
            source_name=f"Trail view {settings.trail_view_id}",
            source_sha256=source_hash,
            metadata={
                "schema_version": "trail-fields-v1",
                "origin": "trail_readonly_snapshot",
                "baseline_scope": settings.baseline_scope,
                "view_id": settings.trail_view_id,
                "fields_visible": list(result.fields_visible),
                "queried_issues": result.queried_issues,
                "returned_issues": result.returned_issues,
                "usable_predictions": len(usable),
                "trigger": trigger,
            },
            rows=snapshot_rows,
            kind="trail_snapshot",
            make_default=False,
            created_by=requested_by,
            created_by_source=identity_source,
            created_by_verified=identity_verified,
        )
        state.update(
            {
                "status": "ready",
                "run_id": run["id"],
                "usable_predictions": len(usable),
                "duplicate": duplicate,
                "message": (
                    result.message
                    + (
                        f" 内容未变化，已复用现有只读快照（{len(usable)} 条）。"
                        if duplicate
                        else f" 已创建只读快照（{len(usable)} 条）。"
                    )
                    + " Trail、GT、人工复核和团队默认 Run 均未修改。"
                ),
            }
        )
        runtime_state["trail_sync"] = state
        return state
    except Exception as exc:
        state = {
            "status": "failed",
            "message": f"Trail 只读{action}失败: {exc}；未创建 Run，团队默认 Run 未变化。",
            "run_id": "",
            "can_create": False,
            "default_changed": False,
            "action": "create" if create_run else "preview",
        }
        runtime_state["trail_sync"] = state
        return state
    finally:
        trail_sync_lock.release()


@asynccontextmanager
async def lifespan(_: FastAPI):
    settings.ensure_directories()
    database.init()
    database.bootstrap_access_users(
        writers=settings.sso_write_users,
        administrators=settings.team_default_managers,
    )
    database.seed_examples(EXAMPLE_CASES)
    bootstrap_baseline()
    asset_index.refresh(force=True)
    bootstrap_model_result()
    if settings.trail_sync_on_start:
        await asyncio.to_thread(
            sync_trail_model_fields,
            create_run=False,
            requested_by="system",
            identity_source="service_startup",
            identity_verified=False,
            trigger="startup",
        )
    batch_prediction_runner.resume_queued_predictions()
    try:
        yield
    finally:
        await asyncio.to_thread(batch_prediction_runner.shutdown)
        await asyncio.to_thread(database.close)


app = FastAPI(
    title="RA Triage Workbench",
    version="1.7.0",
    lifespan=lifespan,
)


@app.middleware("http")
async def identity_ingress_diagnostics(request: Request, call_next):
    if settings.identity_diagnostics and request.url.path.startswith("/api/"):
        candidates = identity_header_candidates(request)
        client_host = request.client.host if request.client else "unknown"
        observation = (client_host, tuple(sorted(candidates.items())))
        if observation not in _identity_diagnostic_observations:
            _identity_diagnostic_observations.add(observation)
            logger.warning(
                "SSO ingress diagnostic client=%s identity_candidates=%s",
                client_host,
                candidates,
            )
    return await call_next(request)


@app.middleware("http")
async def production_write_guard(request: Request, call_next):
    if (
        settings.deployment_mode == "production"
        and request.method in {"POST", "PUT", "PATCH", "DELETE"}
        and request.url.path.startswith("/api/")
    ):
        if not has_same_origin_mutation_marker(request):
            return JSONResponse(
                status_code=403,
                content={
                    "detail": "写请求缺少同源校验标记，请刷新页面后重试。"
                },
            )
        identity = await asyncio.to_thread(request_identity, request, settings)
        role = (
            await asyncio.to_thread(database.access_role, identity.username)
            if identity.verified and identity.username
            else ""
        )
        if role not in {"writer", "admin"}:
            return JSONResponse(
                status_code=403,
                content={
                    "detail": "当前入口为只读预览；请通过获准的企业 SSO 域名访问。"
                },
            )
    return await call_next(request)


@app.middleware("http")
async def request_size_guard(request: Request, call_next):
    if request.method == "POST" and request.url.path in {
        "/api/prediction-batches",
        "/api/import/autotriage",
    }:
        content_length = request.headers.get("content-length", "").strip()
        if not content_length:
            return JSONResponse(
                status_code=411,
                content={"detail": "JSON 请求必须提供 Content-Length。"},
            )
        try:
            request_bytes = int(content_length)
        except ValueError:
            return JSONResponse(
                status_code=400,
                content={"detail": "Content-Length 非法。"},
            )
        if request_bytes > MAX_BATCH_JSON_REQUEST_BYTES:
            return JSONResponse(
                status_code=413,
                content={"detail": "Batch / AutoTriage 请求不能超过 256 KiB。"},
            )
    if (
        request.method == "POST"
        and request.url.path.startswith("/api/cases/")
        and request.url.path.endswith("/annotations-with-attachments")
    ):
        if request.headers.get("x-ra-triage-request") != "review-v1":
            return JSONResponse(
                status_code=403,
                content={"detail": "缺少 Review 截图请求标记。"},
            )
        content_length = request.headers.get("content-length", "").strip()
        if not content_length:
            return JSONResponse(
                status_code=411,
                content={"detail": "Review 截图请求必须提供 Content-Length。"},
            )
        try:
            request_bytes = int(content_length)
        except ValueError:
            return JSONResponse(
                status_code=400,
                content={"detail": "Content-Length 非法。"},
            )
        if request_bytes > MAX_REVIEW_MULTIPART_REQUEST_BYTES:
            return JSONResponse(
                status_code=413,
                content={"detail": "截图上传请求不能超过 26 MB。"},
            )
    return await call_next(request)


@app.middleware("http")
async def request_latency_observer(request: Request, call_next):
    started = time.monotonic()
    response = await call_next(request)
    duration_ms = (time.monotonic() - started) * 1000
    response.headers["Server-Timing"] = f'app;dur={duration_ms:.1f}'
    response.headers["X-Request-Duration-Ms"] = f"{duration_ms:.1f}"
    if request.url.path.startswith("/api/") and duration_ms >= 500:
        logger.warning(
            "slow_request method=%s path=%s status=%s duration_ms=%.1f",
            request.method,
            request.url.path,
            response.status_code,
            duration_ms,
        )
    return response


app.mount("/static", StaticFiles(directory=settings.static_dir), name="static")


@app.get("/", include_in_schema=False)
@app.get("/review", include_in_schema=False)
@app.get("/review-analysis", include_in_schema=False)
@app.get("/runs", include_in_schema=False)
@app.get("/inference", include_in_schema=False)
@app.get("/batch-prediction", include_in_schema=False)
@app.get("/system-status", include_in_schema=False)
@app.get("/users", include_in_schema=False)
async def index() -> HTMLResponse:
    return HTMLResponse(
        content=INDEX_HTML,
        headers={"Cache-Control": "no-store, max-age=0"},
    )


@app.get("/import", include_in_schema=False)
async def legacy_import(kind: str = "issues") -> RedirectResponse:
    # The legacy Issue / GT upload UI is intentionally retired. Keep old
    # bookmarks navigable, but land them in the safe model-result importer.
    return RedirectResponse(
        # Relative Location resolves to /runs for direct IP and
        # /manual/runs through the browser-visible strip-proxy URL.
        url="runs?import=model",
        status_code=307,
    )


@app.get("/health")
async def health() -> dict[str, Any]:
    return {
        "ok": True,
        "build_commit": settings.build_commit,
        "base_path": settings.base_path,
        "deployment_mode": settings.deployment_mode,
        "ra_auto_triage_root_available": (settings.ra_auto_triage_root / "vlm").is_dir(),
        "ares_manifest_available": settings.ares_manifest.is_file(),
        "ares_indexed_issues": asset_index.refresh(),
        "camera_cache_root_available": settings.camera_root.is_dir(),
        "ares_video_root_available": settings.ares_video_root.is_dir(),
        "baseline": runtime_state["baseline"],
        "trail_sync": runtime_state["trail_sync"],
        "trail_write_enabled": False,
        "batch_prediction_enabled": settings.batch_prediction_enabled,
        "autotriage_push_enabled": settings.autotriage_push_enabled,
        "model_gateway": model_catalog.status(),
        "change_revision": await asyncio.to_thread(database.change_revision),
        "storage": database.storage_label,
    }


@app.get("/api/overview")
async def overview(model_run_id: str = "") -> dict[str, Any]:
    selected = model_run_id or await asyncio.to_thread(
        database.default_model_run_id
    )
    return await asyncio.to_thread(
        database.overview,
        baseline_scope=settings.baseline_scope,
        model_run_id=selected,
    )


@app.get("/api/change-revision")
async def change_revision(response: Response) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return {
        "revision": await asyncio.to_thread(database.change_revision),
        "poll_after_ms": 1800,
    }


@app.get("/api/dashboard-config")
async def dashboard_config() -> dict[str, Any]:
    return {
        "baseline": runtime_state["baseline"],
        "build_commit": settings.build_commit,
        "default_model_run_id": database.default_model_run_id(),
        "trail_sync": runtime_state["trail_sync"],
        "missing_evidence_catalog": await asyncio.to_thread(_missing_evidence_catalog),
        "review_tag_catalog": await asyncio.to_thread(_review_tag_catalog),
        # Free-text keyword themes were an earlier experiment and are not
        # exposed by the current structured Review workflow.
        "review_reason_theme_catalog": (),
        "review_attachment_limits": {
            "max_count": MAX_REVIEW_ATTACHMENTS,
            "max_bytes_each": MAX_REVIEW_ATTACHMENT_BYTES,
            "max_bytes_total": MAX_REVIEW_ATTACHMENTS_TOTAL_BYTES,
            "media_types": ["image/png", "image/jpeg", "image/webp"],
        },
        "default_failure_only": bool(database.default_model_run_id()),
        "batch_prediction": {
            "enabled": settings.batch_prediction_enabled,
            "autotriage_push_enabled": settings.autotriage_push_enabled,
            "max_issues": settings.batch_max_issues,
            "input_policy": "server_model_gateway_profile",
            "ares_bev_input": True,
            "trail_write_enabled": False,
            "model_gateway": model_catalog.status(),
        },
    }


def _review_tag_payload(item: dict[str, Any], *, builtin: bool = False) -> dict[str, Any]:
    payload = {
        **item,
        "builtin": bool(item.get("builtin", builtin)),
        "deleted": not bool(item.get("active", 1))
        if "active" in item
        else bool(item.get("deleted", False)),
    }
    payload.pop("active", None)
    payload.setdefault("section", "scene")
    payload.setdefault("group", payload.pop("group_key", "environment"))
    payload.setdefault("hint", "")
    return payload


def _validate_scene_tag_input(body: dict[str, Any]) -> tuple[str, str, str]:
    label = _as_text(body.get("label"))
    hint = _as_text(body.get("hint"))
    group = _as_text(body.get("group") or "environment")
    if not label:
        raise _detail(400, "场景标签标题不能为空。")
    if len(label) > 48 or re.search(r"[\x00-\x1f\x7f]", label):
        raise _detail(400, "场景标签标题长度或字符不合法。")
    if len(hint) > 160 or re.search(r"[\x00-\x1f\x7f]", hint):
        raise _detail(400, "场景标签说明长度或字符不合法。")
    if group not in REVIEW_TAG_SCENE_GROUPS:
        raise _detail(400, "场景标签分组不合法。")
    return label, hint, group


@app.post("/api/review-tags")
async def create_review_tag(request: Request) -> dict[str, Any]:
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "场景标签目录请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "场景标签目录请求必须是 JSON 对象。")
    label, hint, group = _validate_scene_tag_input(body)
    actor, _, _ = _action_actor(request, body.get("created_by"))
    if not actor:
        raise _detail(400, "无法确认场景标签目录创建人。")
    if any(
        str(item.get("label")) == label and not bool(item.get("deleted"))
        for item in _review_tag_catalog()
    ):
        raise _detail(409, "该场景标签标题已经存在。")
    try:
        item = await asyncio.to_thread(
            database.create_review_tag,
            label=label,
            hint=hint,
            section="scene",
            group_key=group,
            created_by=actor,
        )
    except ValueError as exc:
        raise _detail(409, str(exc))
    return {
        "item": _review_tag_payload(item),
        "review_tag_catalog": await asyncio.to_thread(_review_tag_catalog),
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.put("/api/review-tags/{key:path}")
async def update_review_tag(key: str, request: Request) -> dict[str, Any]:
    normalized_key = _as_text(key).strip()
    if not normalized_key or len(normalized_key) > 160:
        raise _detail(400, "场景标签 key 不合法。")
    current = next(
        (item for item in _review_tag_catalog() if item["key"] == normalized_key),
        None,
    )
    if current is None:
        raise _detail(404, "场景标签目录项不存在。")
    if bool(current.get("deleted")):
        raise _detail(409, "该场景标签已经删除。")
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "场景标签目录请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "场景标签目录请求必须是 JSON 对象。")
    section = str(current.get("section") or "scene")
    if section == "scene":
        label, hint, group = _validate_scene_tag_input(body)
    else:
        # Trigger/egress/legacy built-ins keep section/group; only label/hint edit.
        label = _as_text(body.get("label"))
        hint = _as_text(body.get("hint"))
        group = str(current.get("group") or "environment")
        if not label:
            raise _detail(400, "场景标签标题不能为空。")
        if len(label) > 48 or re.search(r"[\x00-\x1f\x7f]", label):
            raise _detail(400, "场景标签标题长度或字符不合法。")
        if len(hint) > 160 or re.search(r"[\x00-\x1f\x7f]", hint):
            raise _detail(400, "场景标签说明长度或字符不合法。")
    actor, _, _ = _action_actor(request, body.get("updated_by"))
    if not actor:
        raise _detail(400, "无法确认场景标签目录编辑人。")
    if any(
        item["key"] != normalized_key
        and str(item.get("label")) == label
        and not bool(item.get("deleted"))
        for item in _review_tag_catalog()
    ):
        raise _detail(409, "该场景标签标题已经存在。")
    try:
        item = await asyncio.to_thread(
            database.update_review_tag,
            key=normalized_key,
            label=label,
            hint=hint,
            section=section,
            group_key=group,
            updated_by=actor,
        )
    except ValueError as exc:
        raise _detail(409, str(exc))
    payload = _review_tag_payload(item, builtin=bool(current.get("builtin")))
    payload["section"] = section
    payload["group"] = group
    return {
        "item": payload,
        "review_tag_catalog": await asyncio.to_thread(_review_tag_catalog),
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.delete("/api/review-tags/{key:path}")
async def delete_review_tag(key: str, request: Request) -> dict[str, Any]:
    normalized_key = _as_text(key).strip()
    if not normalized_key or len(normalized_key) > 160:
        raise _detail(400, "场景标签 key 不合法。")
    current = next(
        (item for item in _review_tag_catalog() if item["key"] == normalized_key),
        None,
    )
    if current is None:
        raise _detail(404, "场景标签目录项不存在。")
    if bool(current.get("deleted")):
        raise _detail(409, "该场景标签已经删除。")
    actor, _, _ = _action_actor(request, "")
    if not actor:
        raise _detail(400, "无法确认场景标签目录删除人。")
    try:
        item = await asyncio.to_thread(
            database.delete_review_tag,
            key=normalized_key,
            deleted_by=actor,
            label=str(current.get("label") or ""),
            hint=str(current.get("hint") or ""),
            section=str(current.get("section") or "scene"),
            group_key=str(current.get("group") or "environment"),
        )
    except ValueError as exc:
        raise _detail(409, str(exc))
    payload = _review_tag_payload(item, builtin=bool(current.get("builtin")))
    payload["deleted"] = True
    return {
        "item": payload,
        "review_tag_catalog": await asyncio.to_thread(_review_tag_catalog),
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.post("/api/missing-evidence")
async def create_missing_evidence(request: Request) -> dict[str, Any]:
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "缺失信息目录请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "缺失信息目录请求必须是 JSON 对象。")
    label = _as_text(body.get("label"))
    hint = _as_text(body.get("hint"))
    actor, _, _ = _action_actor(request, body.get("created_by"))
    if not actor:
        raise _detail(400, "无法确认缺失信息目录创建人。")
    if any(
        str(item["label"]) == label and not bool(item.get("deleted"))
        for item in _missing_evidence_catalog()
    ):
        raise _detail(409, "该缺失信息标题已经存在于内置目录。")
    try:
        item = await asyncio.to_thread(
            database.create_missing_evidence,
            label=label,
            hint=hint,
            created_by=actor,
        )
    except ValueError as exc:
        raise _detail(409, str(exc))
    item = {
        **item,
        "builtin": False,
        "deleted": not bool(item.get("active", 1)),
    }
    item.pop("active", None)
    return {
        "item": item,
        "missing_evidence_catalog": await asyncio.to_thread(_missing_evidence_catalog),
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.put("/api/missing-evidence/{key:path}")
async def update_missing_evidence(key: str, request: Request) -> dict[str, Any]:
    normalized_key = _as_text(key).strip()
    if not normalized_key or len(normalized_key) > 160:
        raise _detail(400, "缺失信息 key 不合法。")
    current = next(
        (item for item in _missing_evidence_catalog() if item["key"] == normalized_key),
        None,
    )
    if current is None:
        raise _detail(404, "缺失信息目录项不存在。")
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "缺失信息目录请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "缺失信息目录请求必须是 JSON 对象。")
    label = _as_text(body.get("label"))
    hint = _as_text(body.get("hint"))
    actor, _, _ = _action_actor(request, body.get("updated_by"))
    if not actor:
        raise _detail(400, "无法确认缺失信息目录编辑人。")
    if any(
        item["key"] != normalized_key
        and not bool(item.get("deleted"))
        and str(item["label"]) == label
        for item in _missing_evidence_catalog()
    ):
        raise _detail(409, "该缺失信息标题已经存在。")
    try:
        item = await asyncio.to_thread(
            database.update_missing_evidence,
            key=normalized_key,
            label=label,
            hint=hint,
            updated_by=actor,
        )
    except ValueError as exc:
        raise _detail(409, str(exc))
    item = {
        **item,
        "builtin": bool(current.get("builtin")),
        "deleted": not bool(item.get("active", 1)),
    }
    item.pop("active", None)
    return {
        "item": item,
        "missing_evidence_catalog": await asyncio.to_thread(_missing_evidence_catalog),
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.delete("/api/missing-evidence/{key:path}")
async def delete_missing_evidence(key: str, request: Request) -> dict[str, Any]:
    normalized_key = _as_text(key).strip()
    if not normalized_key or len(normalized_key) > 160:
        raise _detail(400, "缺失信息 key 不合法。")
    current = next(
        (item for item in _missing_evidence_catalog() if item["key"] == normalized_key),
        None,
    )
    if current is None:
        raise _detail(404, "缺失信息目录项不存在。")
    if bool(current.get("deleted")):
        raise _detail(409, "该缺失信息已经删除。")
    try:
        body = await request.json()
    except (TypeError, ValueError):
        body = {}
    if not isinstance(body, dict):
        body = {}
    actor, _, _ = _action_actor(request, body.get("deleted_by"))
    if not actor:
        raise _detail(400, "无法确认缺失信息目录删除人。")
    try:
        item = await asyncio.to_thread(
            database.delete_missing_evidence,
            key=normalized_key,
            label=str(current.get("label") or ""),
            hint=str(current.get("hint") or ""),
            deleted_by=actor,
        )
    except ValueError as exc:
        raise _detail(409, str(exc))
    item = {
        **item,
        "builtin": bool(current.get("builtin")),
        "deleted": True,
    }
    item.pop("active", None)
    return {
        "item": item,
        "missing_evidence_catalog": await asyncio.to_thread(_missing_evidence_catalog),
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.get("/api/session")
async def session(request: Request, response: Response) -> dict[str, object]:
    response.headers["Cache-Control"] = "no-store, max-age=0"
    identity = await asyncio.to_thread(request_identity, request, settings)
    access_role = (
        await asyncio.to_thread(database.access_role, identity.username)
        if identity.verified and identity.username
        else ""
    )
    is_admin = access_role == "admin"
    can_write = settings.deployment_mode != "production" or access_role in {
        "writer",
        "admin",
    }
    payload = identity.as_dict(
        trust_proxy_headers=settings.trust_proxy_identity_headers
    ) | {
        "browser_lca_fallback": not identity.authenticated,
        "identity_header": (
            settings.identity_header
            if settings.trust_proxy_identity_headers
            else ""
        ),
        "access_role": access_role or "viewer",
        "is_admin": is_admin,
        "can_manage_team_default": is_admin,
        "deployment_mode": settings.deployment_mode,
        "can_write": can_write,
        "read_only": not can_write,
        "login_managed_by": "kylin" if settings.kylin_sso_enabled else "",
        "logout_url": with_base_path(settings.base_path, "/auth/logout")
        if settings.kylin_sso_enabled
        else "",
    }
    if settings.identity_diagnostics:
        payload["identity_header_candidates"] = identity_header_candidates(request)
    return payload


@app.get("/api/access-users")
async def list_access_users(request: Request) -> dict[str, Any]:
    await asyncio.to_thread(_admin_identity, request)
    return {"items": await asyncio.to_thread(database.list_access_users)}


@app.put("/api/access-users/{username}")
async def set_access_user(
    username: str, request: Request
) -> dict[str, Any]:
    identity = await asyncio.to_thread(_admin_identity, request)
    normalized = normalise_username(username)
    if not normalized:
        raise _detail(400, "用户名格式不合法。")
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "请求 JSON 不合法。")
    role = _as_text(body.get("role") if isinstance(body, dict) else "").lower()
    try:
        user = await asyncio.to_thread(
            database.set_access_user,
            username=normalized,
            role=role,
            actor=identity.username,
        )
    except ValueError as exc:
        raise _detail(409, str(exc))
    return {
        "user": user,
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.delete("/api/access-users/{username}")
async def delete_access_user(username: str, request: Request) -> dict[str, Any]:
    await asyncio.to_thread(_admin_identity, request)
    normalized = normalise_username(username)
    if not normalized:
        raise _detail(400, "用户名格式不合法。")
    try:
        deleted = await asyncio.to_thread(database.delete_access_user, normalized)
    except ValueError as exc:
        raise _detail(409, str(exc))
    if not deleted:
        raise _detail(404, "用户权限记录不存在。")
    return {
        "deleted": True,
        "username": normalized,
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.get("/auth/logout")
async def logout() -> RedirectResponse:
    response = RedirectResponse(settings.kylin_sso_logout_url, status_code=302)
    for cookie_name in ("_kylin_ticket", "_kylin_username"):
        response.delete_cookie(cookie_name, path="/")
    return response


@app.get("/api/status")
async def status(response: Response) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, max-age=0"
    try:
        database_state = await asyncio.to_thread(
            database.runtime_status,
            persistent_data=settings.postgres_persistent_data,
        )
    except Exception:
        logger.exception("database status check failed")
        database_state = {
            "ok": False,
            "backend": database.backend,
            "server_version": "",
            "persistent_data": False,
            "revision": 0,
            "migration_count": 0,
            "pool_max_size": database.pool_size,
            "latency_ms": None,
        }
    backup_state, volume_state = await asyncio.gather(
        asyncio.to_thread(backup_status, settings.data_dir),
        asyncio.to_thread(volume_status, settings.data_dir),
    )
    baseline_state = runtime_state["baseline"]
    overall = overall_status(
        database=database_state,
        baseline=baseline_state,
        backups=backup_state,
        volume=volume_state,
    )
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "base_path": settings.base_path,
        "overall": overall,
        "application": {
            "started_at": APP_STARTED_AT.isoformat(),
            "uptime_seconds": max(0, int(time.monotonic() - APP_STARTED_MONOTONIC)),
            "build_commit": settings.build_commit,
            "base_path": settings.base_path,
            "deployment_mode": settings.deployment_mode,
        },
        "trail_write_enabled": False,
        "build_commit": settings.build_commit,
        "ra_auto_triage_root_available": (settings.ra_auto_triage_root / "vlm").is_dir(),
        "ares_manifest_available": settings.ares_manifest.is_file(),
        "ares_indexed_issues": asset_index.refresh(),
        "camera_cache_root_available": settings.camera_root.is_dir(),
        "ares_video_root_available": settings.ares_video_root.is_dir(),
        "baseline": baseline_state,
        "trail_sync": runtime_state["trail_sync"],
        "batch_prediction_enabled": settings.batch_prediction_enabled,
        "autotriage_push_enabled": settings.autotriage_push_enabled,
        "batch_max_issues": settings.batch_max_issues,
        "model_gateway": model_catalog.status(),
        "storage": database.storage_label,
        "database": database_state,
        "backups": backup_state,
        "volume": volume_state,
        "model_endpoint_policy": (
            "固定服务器网关；Profile 模型已验证，其他在线 Qwen3 显式实验；"
            "浏览器不能提交地址或凭证"
        ),
    }


def _parse_issue_id_filter(raw: str) -> list[str]:
    tokens = re.split(r"[\s,;|]+", _as_text(raw))
    cleaned: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        issue_id = token.strip()
        if not issue_id or issue_id in seen:
            continue
        if not re.fullmatch(r"[A-Za-z0-9_-]{3,128}", issue_id):
            continue
        seen.add(issue_id)
        cleaned.append(issue_id)
        if len(cleaned) >= 2000:
            break
    return cleaned


def _case_filter_kwargs(
    *,
    search: str = "",
    gt_label: str = "",
    model_label: str = "",
    annotation_label: str = "",
    annotation_author: str = "",
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    missing_evidence: str = "",
    issue_ids: str = "",
    work_assignee: str = "",
) -> dict[str, Any]:
    comparison_values = [
        value
        for value in _csv_filter_values(comparison)
        if value in COMPARISON_STATUSES and value != "all"
    ]
    if failure_only and comparison_values and comparison_values != ["mismatch"]:
        # Legacy failure_only=true only expands empty comparison to mismatch.
        if comparison and set(comparison_values) != {"mismatch"}:
            raise _detail(400, "failure_only=true 与 comparison 参数冲突。")
    if failure_only and not comparison_values:
        comparison_values = ["mismatch"]
    if comparison_values and set(comparison_values) == {
        "match",
        "mismatch",
        "none",
    }:
        comparison_values = []
    comparison_status = ",".join(comparison_values) if comparison_values else "all"
    if comparison_status != "all" and not model_run_id:
        raise _detail(400, "筛选模型对比关系时必须选择 Model Run。")
    model_labels = _csv_filter_values(model_label)
    for label in model_labels:
        if label not in LABELS:
            raise _detail(400, "model_label 不在三分类范围内。")
    if model_labels and not model_run_id:
        raise _detail(400, "按模型标注筛选时必须选择 Model Run。")
    gt_labels = _csv_filter_values(gt_label)
    for label in gt_labels:
        if label not in LABELS:
            raise _detail(400, "gt_label 不在三分类范围内。")
    return {
        "baseline_scope": settings.baseline_scope,
        "search": search,
        "gt_label": ",".join(gt_labels),
        "model_label": ",".join(model_labels),
        "annotation_label": annotation_label,
        "annotation_author": annotation_author,
        "model_run_id": model_run_id,
        "comparison_status": comparison_status,
        "failure_only": failure_only,
        "missing_evidence": missing_evidence,
        "issue_ids": _parse_issue_id_filter(issue_ids),
        "work_assignee": _as_text(work_assignee).strip(),
    }


@app.get("/api/cases")
async def list_cases(
    search: str = "",
    gt_label: str = "",
    model_label: str = "",
    annotation_label: str = "",
    annotation_author: str = "",
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    missing_evidence: str = "",
    issue_ids: str = "",
    work_assignee: str = "",
    page: int = 1,
    page_size: int = 100,
    include_thumbnail: bool = False,
) -> dict[str, Any]:
    filters = _case_filter_kwargs(
        search=search,
        gt_label=gt_label,
        model_label=model_label,
        annotation_label=annotation_label,
        annotation_author=annotation_author,
        model_run_id=model_run_id,
        comparison=comparison,
        failure_only=failure_only,
        missing_evidence=missing_evidence,
        issue_ids=issue_ids,
        work_assignee=work_assignee,
    )
    comparison_status = filters["comparison_status"]
    result = await asyncio.to_thread(
        database.list_cases,
        **filters,
        page=page,
        page_size=min(max(1, int(page_size)), 100),
    )
    items: list[dict[str, Any]] = []
    for item in result.get("items", []):
        issue_id = _as_text(item.get("issue_id"))
        public = {
            **item,
            "voyager_issue_url": _voyager_issue_url(issue_id),
        }
        if include_thumbnail:
            public["thumbnail"] = (
                {
                    "url": _public_path(
                        f"/api/case-thumbnails/{quote(issue_id, safe='')}"
                    ),
                    "kind": "bev",
                    "label": "BEV · t0 附近",
                }
                if asset_index.has_issue(issue_id)
                else None
            )
        items.append(public)
    result["items"] = items
    result["filters"] = {
        "model_run_id": model_run_id,
        "comparison_status": comparison_status,
        "failure_only": comparison_status == "mismatch",
        "issue_ids": filters["issue_ids"],
        "work_assignee": filters["work_assignee"],
    }
    return result


@app.get("/api/cases/issue-ids")
async def list_case_issue_ids(
    search: str = "",
    gt_label: str = "",
    model_label: str = "",
    annotation_label: str = "",
    annotation_author: str = "",
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    missing_evidence: str = "",
    issue_ids: str = "",
    work_assignee: str = "",
) -> dict[str, Any]:
    """Return all matching issue IDs for the current Review filters (capped)."""

    filters = _case_filter_kwargs(
        search=search,
        gt_label=gt_label,
        model_label=model_label,
        annotation_label=annotation_label,
        annotation_author=annotation_author,
        model_run_id=model_run_id,
        comparison=comparison,
        failure_only=failure_only,
        missing_evidence=missing_evidence,
        issue_ids=issue_ids,
        work_assignee=work_assignee,
    )
    ids = await asyncio.to_thread(database.list_case_issue_ids, **filters, limit=5000)
    return {
        "issue_ids": ids,
        "total": len(ids),
        "truncated": len(ids) >= 5000,
        "filters": {
            "model_run_id": model_run_id,
            "comparison_status": filters["comparison_status"],
            "failure_only": filters["comparison_status"] == "mismatch",
            "issue_ids": filters["issue_ids"],
            "work_assignee": filters["work_assignee"],
        },
    }


@app.get("/api/work-assignees")
async def work_assignees() -> dict[str, Any]:
    """Assignees currently attached to Issues via work-split."""

    return {
        "items": await asyncio.to_thread(database.list_work_assignees),
    }


@app.post("/api/cases/work-split")
async def split_case_work(request: Request) -> dict[str, Any]:
    """Admin-only: randomly assign filtered issues and persist ownership."""

    identity = _admin_identity(request)
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "均分任务请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "均分任务请求必须是 JSON 对象。")
    filter_body = body.get("filters") if isinstance(body.get("filters"), dict) else {}
    filters = _case_filter_kwargs(
        search=_as_text(filter_body.get("search")),
        gt_label=_as_text(filter_body.get("gt_label")),
        model_label=_as_text(filter_body.get("model_label")),
        annotation_label=_as_text(filter_body.get("annotation_label")),
        annotation_author=_as_text(filter_body.get("annotation_author")),
        model_run_id=_as_text(filter_body.get("model_run_id")),
        comparison=_as_text(filter_body.get("comparison") or filter_body.get("comparison_status")),
        failure_only=bool(filter_body.get("failure_only")),
        missing_evidence=_as_text(filter_body.get("missing_evidence")),
        issue_ids=_as_text(filter_body.get("issue_ids")),
        work_assignee=_as_text(filter_body.get("work_assignee")),
    )
    issue_ids = await asyncio.to_thread(database.list_case_issue_ids, **filters, limit=5000)
    assignees = body.get("assignees")
    if not isinstance(assignees, list):
        raise _detail(400, "assignees 必须是数组。")
    raw_seed = body.get("seed", None)
    seed: int | None
    if raw_seed in (None, ""):
        seed = None
    else:
        try:
            seed = int(raw_seed)
        except (TypeError, ValueError):
            raise _detail(400, "seed 必须是整数。")
    try:
        assignments = distribute_issue_ids(issue_ids, assignees, seed=seed)
        saved = await asyncio.to_thread(
            database.apply_work_split,
            assignments=assignments,
            created_by=identity.username,
            seed=seed,
            filter_snapshot={
                "model_run_id": filters["model_run_id"],
                "comparison_status": filters["comparison_status"],
                "search": filters["search"],
                "gt_label": filters["gt_label"],
                "model_label": filters["model_label"],
                "annotation_author": filters["annotation_author"],
                "missing_evidence": filters["missing_evidence"],
            },
        )
    except ValueError as exc:
        raise _detail(400, str(exc))
    return {
        "total": len(issue_ids),
        "truncated": len(issue_ids) >= 5000,
        "seed": seed,
        "split_id": saved["split_id"],
        "created_by": saved["created_by"],
        "created_at": saved["created_at"],
        "assignments": assignments,
        "work_assignees": await asyncio.to_thread(database.list_work_assignees),
        "change_revision": await asyncio.to_thread(database.change_revision),
        "filters": {
            "model_run_id": filters["model_run_id"],
            "comparison_status": filters["comparison_status"],
            "failure_only": filters["comparison_status"] == "mismatch",
            "work_assignee": filters["work_assignee"],
        },
    }


@app.get("/api/reviewers")
async def reviewers(model_run_id: str = "") -> dict[str, Any]:
    return {
        "items": await asyncio.to_thread(
            database.list_reviewers, settings.baseline_scope, model_run_id
        )
    }


@app.get("/api/case-thumbnails/{issue_id}")
async def get_case_thumbnail(issue_id: str) -> FileResponse:
    if not ISSUE_ID_RE.fullmatch(issue_id) or await asyncio.to_thread(
        database.get_case, issue_id
    ) is None:
        raise _detail(404, "Issue 不存在。")
    source_info = await asyncio.to_thread(
        asset_index.get_thumbnail_source,
        issue_id,
    )
    if not source_info or not isinstance(source_info.get("path"), Path):
        raise _detail(404, "该 Issue 暂无 BEV 缩略图。")
    source = source_info["path"]
    try:
        destination = _thumbnail_cache_path(issue_id, source)
        if not destination.is_file():
            async with thumbnail_image_semaphore:
                if not destination.is_file():
                    await asyncio.to_thread(
                        _render_case_thumbnail,
                        source,
                        destination,
                    )
    except (
        Image.DecompressionBombError,
        UnidentifiedImageError,
        OSError,
        ValueError,
    ):
        raise _detail(404, "该 Issue 的 BEV 缩略图无法生成。")
    return FileResponse(
        destination,
        media_type="image/jpeg",
        headers={"Cache-Control": "public, max-age=300, must-revalidate"},
    )


@app.get("/api/cases/{issue_id}/trail-metadata")
async def get_case_trail_metadata(issue_id: str) -> dict[str, Any]:
    """Load optional Trail metadata without delaying the Issue detail API.

    The Issue detail response intentionally exposes only local data and any
    metadata already imported with the case.  Trail is a best-effort external
    dependency and is fetched by the browser after the detail has rendered.
    """

    case = await asyncio.to_thread(database.get_case, issue_id)
    if case is None:
        raise _detail(404, "Issue 不存在。")

    trail_metadata = _case_link_metadata_fallback(case)
    status = "disabled"
    if settings.trail_detail_metadata_enabled:
        status = "unavailable"
        try:
            async with trail_detail_semaphore:
                fetched = await asyncio.wait_for(
                    asyncio.to_thread(
                        read_trail_issue_metadata,
                        ra_root=settings.ra_auto_triage_root,
                        issue_id=issue_id,
                        view_id=settings.trail_view_id,
                        cache_seconds=settings.trail_detail_metadata_cache_seconds,
                    ),
                    timeout=8.0,
                )
            trail_metadata.update(fetched)
            status = "ready" if fetched else "unavailable"
        except asyncio.TimeoutError:
            status = "timeout"
            logger.warning("Trail detail metadata timed out issue_id=%s", issue_id)
        except Exception:
            status = "unavailable"
            logger.warning("Trail detail metadata unavailable issue_id=%s", issue_id)

    return {
        "issue_id": issue_id,
        "status": status,
        "external_links": _case_external_links(issue_id, trail_metadata),
    }


@app.get("/api/cases/{issue_id}")
async def get_case(issue_id: str) -> dict[str, Any]:
    case = await asyncio.to_thread(database.get_case, issue_id)
    if case is None:
        raise _detail(404, "Issue 不存在。")
    for annotation in case.get("annotations", []):
        annotation["attachments"] = [
            _public_review_attachment(attachment)
            for attachment in annotation.get("attachments", [])
        ]
    case["assets"] = asset_index.get_assets(issue_id)
    captured_video = video_index.get_video(issue_id)
    if captured_video is not None:
        case["assets"]["video"] = captured_video
        case["assets"]["available"] = True
    case["camera"] = camera_index.get_assets(
        issue_id, (case["assets"].get("capture") or {}).get("timestamp_ms")
    )
    case["voyager_issue_url"] = _voyager_issue_url(issue_id)
    case["external_links"] = _case_external_links(
        issue_id, _case_link_metadata_fallback(case)
    )
    case["trail_metadata_status"] = (
        "pending" if settings.trail_detail_metadata_enabled else "disabled"
    )
    case["batch_jobs"] = [
        _public_batch_job(job) for job in case.get("batch_jobs", [])
    ]
    return case


@app.get("/api/assets/{issue_id}/{asset_id}")
async def get_asset(issue_id: str, asset_id: str) -> FileResponse:
    path = (
        asset_index.get_asset_path(issue_id, asset_id)
        or camera_index.get_asset_path(issue_id, asset_id)
        or video_index.get_asset_path(issue_id, asset_id)
    )
    if path is None:
        raise _detail(404, "Ares / Camera 资产不存在。")
    suffix = path.suffix.lower()
    media_type = (
        "video/mp4"
        if suffix == ".mp4"
        else "image/jpeg"
        if suffix in {".jpg", ".jpeg"}
        else "image/png"
    )
    if suffix == ".mp4":
        return FileResponse(
            path,
            media_type=media_type,
            headers={
                "Accept-Ranges": "bytes",
                "Cache-Control": "private, max-age=300, must-revalidate",
                "Content-Disposition": "inline",
                "X-Content-Type-Options": "nosniff",
            },
        )
    return FileResponse(path, media_type=media_type, filename=path.name)


@app.get("/api/review-attachments/{attachment_id}")
async def get_review_attachment(attachment_id: str) -> FileResponse:
    attachment = database.get_review_attachment(attachment_id)
    if attachment is None:
        raise _detail(404, "Review 截图不存在。")
    root = settings.review_attachments_dir.resolve()
    path = (root / attachment["stored_name"]).resolve()
    if root not in path.parents or not path.is_file():
        raise _detail(404, "Review 截图文件不存在。")
    return FileResponse(
        path,
        media_type=attachment["media_type"],
        headers={
            "Content-Disposition": "inline",
            "X-Content-Type-Options": "nosniff",
            "Cache-Control": "private, max-age=31536000, immutable",
        },
    )


@app.post("/api/cases/{issue_id}/annotations")
async def create_annotation(issue_id: str, request: Request) -> dict[str, Any]:
    if database.get_case(issue_id) is None:
        raise _detail(404, "Issue 不存在。")
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "标注请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "标注请求必须是 JSON 对象。")
    annotation = _create_annotation_record(
        issue_id=issue_id,
        request=request,
        body=body,
    )
    return {
        "annotation": annotation,
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.post("/api/cases/{issue_id}/annotations-with-attachments")
async def create_annotation_with_attachments(
    issue_id: str,
    request: Request,
    payload: str = Form(...),
    attachments: list[UploadFile] | None = File(None),
) -> dict[str, Any]:
    if database.get_case(issue_id) is None:
        raise _detail(404, "Issue 不存在。")
    try:
        body = json.loads(payload)
    except (TypeError, ValueError, json.JSONDecodeError):
        raise _detail(400, "payload 必须是 JSON 对象。")
    if not isinstance(body, dict):
        raise _detail(400, "payload 必须是 JSON 对象。")
    records, paths = await _store_review_attachments(attachments or [])
    try:
        annotation = _create_annotation_record(
            issue_id=issue_id,
            request=request,
            body=body,
            attachments=records,
        )
    except Exception:
        for path in paths:
            path.unlink(missing_ok=True)
        raise
    annotation["attachments"] = [
        _public_review_attachment(attachment)
        for attachment in annotation.get("attachments", [])
    ]
    return {
        "annotation": annotation,
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.delete("/api/cases/{issue_id}/annotations/{annotation_id}")
async def delete_annotation(issue_id: str, annotation_id: int) -> dict[str, Any]:
    if annotation_id <= 0:
        raise _detail(400, "Review 版本 ID 不合法。")
    deleted = await asyncio.to_thread(
        database.delete_annotation,
        issue_id=issue_id,
        annotation_id=annotation_id,
    )
    if deleted is None:
        raise _detail(404, "Review 版本不存在或已被删除。")
    attachment_root = settings.review_attachments_dir.resolve()
    for attachment in deleted.get("attachments", []):
        path = (attachment_root / str(attachment.get("stored_name") or "")).resolve()
        if attachment_root not in path.parents:
            logger.warning(
                "Skipped unsafe deleted review attachment path annotation_id=%s",
                annotation_id,
            )
            continue
        try:
            path.unlink(missing_ok=True)
        except OSError:
            logger.exception(
                "Failed to remove deleted review attachment annotation_id=%s attachment_id=%s",
                annotation_id,
                attachment.get("id"),
            )
    deleted["attachments"] = [
        _public_review_attachment(attachment)
        for attachment in deleted.get("attachments", [])
    ]
    return {
        "deleted": deleted,
        "change_revision": await asyncio.to_thread(database.change_revision),
    }


@app.get("/api/model-runs")
async def model_runs() -> dict[str, Any]:
    items = await asyncio.to_thread(
        database.list_model_runs, settings.baseline_scope
    )
    for item in items:
        if item.get("kind") != "upload":
            continue
        source_file = _model_source_file(item)
        filename = _model_source_filename(item)
        suffix = Path(filename).suffix.lower()
        preview_supported = suffix in {".json", ".csv", ".xlsx", ".xlsm"}
        reconstructed = (
            source_file is None
            and preview_supported
            and bool(item.get("prediction_count"))
        )
        item["source_file"] = {
            "filename": filename,
            "available": bool(source_file) or reconstructed,
            "reconstructed": reconstructed,
            "preview_supported": preview_supported,
            "preview_url": _public_path(
                f"/api/model-runs/{quote(str(item['id']), safe='')}/source-preview"
                if preview_supported
                else f"/api/model-runs/{quote(str(item['id']), safe='')}/source"
            ),
            "download_url": _public_path(
                f"/api/model-runs/{quote(str(item['id']), safe='')}/source?download=1"
            ),
        }
    return {
        "items": items,
        "default_model_run_id": database.default_model_run_id(),
    }


@app.get("/api/model-runs/{run_id}/source-preview")
async def preview_model_run_source(
    run_id: str,
    page: int = 1,
    page_size: int = 100,
) -> dict[str, Any]:
    run = database.get_model_run(run_id)
    if run is None:
        raise _detail(404, "模型 Run 不存在。")
    source_file = _model_source_file(run)
    reconstructed = False
    if source_file is None:
        filename = _model_source_filename(run)
        suffix = Path(filename).suffix.lower()
        reconstructed = True
    else:
        path, filename = source_file
        suffix = path.suffix.lower()
    if suffix not in {".json", ".csv", ".xlsx", ".xlsm"}:
        raise _detail(415, "当前仅支持在页面内预览 CSV / JSON / XLSX。")
    if reconstructed:
        source_rows = database.model_run_source_rows(run_id)
        metadata = {
            "source": "dashboard_reconstructed",
            "notice": "原始文件未归档；以下内容由该 Run 已保存的脱敏预测行重建。",
        }
    else:
        try:
            if path.stat().st_size > MAX_UPLOAD_BYTES:
                raise _detail(413, "来源文件过大，暂不生成页面预览；请使用下载。")
            content = await asyncio.to_thread(path.read_bytes)
            source_rows, metadata = parse_source_bytes(filename, content)
        except HTTPException:
            raise
        except (OSError, UnicodeDecodeError, ValueError, json.JSONDecodeError) as exc:
            raise _detail(422, f"来源文件无法生成预览：{exc}")

    page = max(1, int(page))
    page_size = min(max(1, int(page_size)), MAX_SOURCE_PREVIEW_ROWS)
    total_rows = len(source_rows)
    page_count = max(1, math.ceil(total_rows / page_size))
    page = min(page, page_count)
    page_start = (page - 1) * page_size
    page_rows = source_rows[page_start : page_start + page_size]
    preview_rows = [
        {
            str(key): _source_preview_value(value)
            for key, value in row.items()
        }
        for row in page_rows
    ]
    columns: list[str] = []
    seen_columns: set[str] = set()
    for row in preview_rows:
        for key in row:
            if key not in seen_columns:
                seen_columns.add(key)
                columns.append(key)
    metadata_preview = {
        str(key): _source_preview_value(value)
        for key, value in (metadata.items() if isinstance(metadata, dict) else [])
    }
    return {
        "filename": filename,
        "format": suffix[1:],
        "columns": columns,
        "rows": preview_rows,
        "total_rows": total_rows,
        "page": page,
        "page_size": page_size,
        "page_count": page_count,
        "offset": page_start,
        "has_previous": page > 1,
        "has_next": page < page_count,
        "truncated": total_rows > len(preview_rows),
        "metadata": metadata_preview,
        "reconstructed": reconstructed,
    }


@app.get("/api/model-runs/{run_id}/source")
async def get_model_run_source(run_id: str, download: bool = False) -> Response:
    run = database.get_model_run(run_id)
    if run is None:
        raise _detail(404, "模型 Run 不存在。")
    source_file = _model_source_file(run)
    if source_file is None:
        reconstructed = _reconstructed_model_source(run_id, run)
        if reconstructed is None:
            raise _detail(404, "该 Run 的原始文件未归档且无法从已保存结果重建。")
        content, filename, media_type = reconstructed
        disposition = "attachment" if download else "inline"
        return Response(
            content=content,
            media_type=media_type,
            headers={
                "Content-Disposition": f'{disposition}; filename="{filename}"',
                "Cache-Control": "private, max-age=300, must-revalidate",
                "X-Content-Type-Options": "nosniff",
                "X-RA-Source-Reconstructed": "1",
            },
        )
    path, filename = source_file
    media_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    disposition = "attachment" if download else "inline"
    return FileResponse(
        path,
        media_type=media_type,
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"',
            "Cache-Control": "private, max-age=300, must-revalidate",
            "X-Content-Type-Options": "nosniff",
        },
    )


@app.delete("/api/model-runs/{run_id}")
async def delete_model_run(run_id: str) -> dict[str, Any]:
    run = database.get_model_run(run_id)
    if run is None:
        raise _detail(404, "模型 Run 不存在。")
    source_file = _model_source_file(run)
    try:
        deleted = database.delete_model_run(run_id)
    except ValueError as exc:
        raise _detail(409, str(exc))
    if deleted is None:
        raise _detail(404, "模型 Run 不存在。")

    source_deleted = False
    if source_file is not None:
        path, _ = source_file
        upload_root = settings.uploads_dir.resolve()
        resolved = path.resolve()
        if upload_root == resolved or upload_root in resolved.parents:
            try:
                resolved.unlink(missing_ok=True)
                source_deleted = True
            except OSError:
                # The Run is already deleted; report the artifact separately so
                # an operator can recover a permissions/disk cleanup issue.
                source_deleted = False
    return {
        "deleted": deleted,
        "source_deleted": source_deleted,
    }


@app.post("/api/model-runs/{run_id}/default")
async def set_default_model_run(run_id: str, request: Request) -> dict[str, Any]:
    if not _can_manage_team_default(request):
        raise _detail(
            403,
            "设置团队默认 Run 需要可信 SSO 且用户名位于 "
            "DASHBOARD_TEAM_DEFAULT_MANAGERS；当前仍可在 Review 中选择任意 Run。",
        )
    run = database.set_default_model_run(run_id)
    if run is None:
        raise _detail(404, "模型 run 不存在。")
    return {"run": run, "default_model_run_id": run_id}


@app.get("/api/review-clusters")
async def review_clusters(
    model_run_id: str = "",
    failure_only: bool = True,
    annotation_author: str = "",
) -> dict[str, Any]:
    return {
        "items": database.review_clusters(
            baseline_scope=settings.baseline_scope,
            model_run_id=model_run_id,
            failure_only=failure_only,
            annotation_author=annotation_author,
        )
    }


def _csv_filter_values(raw: str | list[str] | tuple[str, ...] | None) -> tuple[str, ...]:
    values: list[str] = []
    if raw is None:
        return ()
    if isinstance(raw, (list, tuple)):
        parts = raw
    else:
        parts = str(raw).split(",")
    for part in parts:
        text = _as_text(part).strip()
        if text:
            values.append(text)
    return tuple(dict.fromkeys(values))


def _review_reason_analysis_payload(
    *,
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    annotation_author: str = "",
    review_status: str = "",
    gt_label: str = "",
    annotation_label: str = "",
    model_label: str = "",
    missing_evidence: str = "",
    theme: str = "",
    tag: str = "",
    scene_tag: str = "",
    trigger_tag: str = "",
    egress_tag: str = "",
    search: str = "",
    page: int = 1,
    page_size: int = 20,
    unbounded: bool = False,
) -> dict[str, Any]:
    missing_evidence_catalog = _missing_evidence_catalog()
    evidence_catalog = {
        str(item["key"]): {
            "label": str(item["label"]),
            "description": str(item["hint"]),
        }
        for item in missing_evidence_catalog
    }
    # Free-text keyword themes are not part of the current Review contract.
    # Ignore the retired theme parameter so old bookmarked URLs remain usable;
    # the canonical response and UI no longer expose or apply it.
    theme = ""
    requested_comparison = _as_text(comparison).strip().lower()
    if requested_comparison and requested_comparison not in COMPARISON_STATUSES:
        raise _detail(
            400,
            "comparison 仅支持 all、mismatch、match 或 none。",
        )
    if (
        failure_only
        and requested_comparison
        and requested_comparison != "mismatch"
    ):
        raise _detail(400, "failure_only=true 与 comparison 参数冲突。")
    comparison_status = requested_comparison or (
        "mismatch" if failure_only else "all"
    )
    authors = _csv_filter_values(annotation_author)
    statuses = _csv_filter_values(review_status)
    gt_labels = _csv_filter_values(gt_label)
    annotation_labels = _csv_filter_values(annotation_label)
    model_labels = _csv_filter_values(model_label)
    evidence_keys = _csv_filter_values(missing_evidence)
    for status in statuses:
        if status not in REVIEW_STATUSES:
            raise _detail(400, "review_status 不在支持范围内。")
    for label in gt_labels:
        if label not in LABELS:
            raise _detail(400, "gt_label 不在三分类范围内。")
    for label in annotation_labels:
        if label not in LABELS:
            raise _detail(400, "annotation_label 不在三分类范围内。")
    for label in model_labels:
        if label not in LABELS:
            raise _detail(400, "model_label 不在三分类范围内。")
    if model_labels and not model_run_id:
        raise _detail(400, "按模型预测筛选时必须选择 Model Run。")
    for key in evidence_keys:
        if key not in evidence_catalog:
            raise _detail(400, "missing_evidence 不在稳定字段目录中。")
    tag_catalog = _review_tag_catalog()
    tag_by_key = {str(item["key"]): item for item in tag_catalog}
    scene_tags = _csv_filter_values(scene_tag)
    trigger_tags = _csv_filter_values(trigger_tag)
    egress_tags = _csv_filter_values(egress_tag)
    legacy_tags = _csv_filter_values(tag)
    for requested_tag in (*legacy_tags, *scene_tags, *trigger_tags, *egress_tags):
        item = tag_by_key.get(requested_tag)
        if item is None:
            raise _detail(400, "场景 Tags 不在共享目录中。")
    for key in scene_tags:
        if tag_by_key[key].get("section") != "scene":
            raise _detail(400, "scene_tag 必须属于场景 Tags。")
    for key in trigger_tags:
        if tag_by_key[key].get("section") != "interaction_decision":
            raise _detail(400, "trigger_tag 必须属于触发判定 Tags。")
    for key in egress_tags:
        if tag_by_key[key].get("section") != "egress":
            raise _detail(400, "egress_tag 必须属于如何脱困 Tags。")
    if comparison_status != "all" and not model_run_id:
        raise _detail(400, "筛选模型对比关系时必须选择 Model Run。")

    selected_run: dict[str, Any] | None = None
    if model_run_id:
        selected_run = next(
            (
                run
                for run in database.list_model_runs(settings.baseline_scope)
                if run["id"] == model_run_id
            ),
            None,
        )
        if selected_run is None:
            raise _detail(404, "Model Run 不存在。")

    normalized_search = _as_text(search)[:256]
    folded_search = normalized_search.casefold()
    search_aliases = tuple(
        str(item["key"])
        for item in (*_review_tag_catalog(), *missing_evidence_catalog)
        if folded_search
        and (
            folded_search in str(item["label"]).casefold()
            or str(item["label"]).casefold() in folded_search
        )
    )
    status_labels = {
        "待复核": "pending",
        "已 Review": "reviewed",
        "GT 待复核": "needs_gt_review",
    }
    search_aliases += tuple(
        status
        for label, status in status_labels.items()
        if folded_search
        and (
            folded_search in label.casefold()
            or label.casefold() in folded_search
        )
    )
    rows = database.review_reason_rows(
        baseline_scope=settings.baseline_scope,
        model_run_id=model_run_id,
        comparison_status=comparison_status,
        annotation_author=",".join(authors),
        review_status=",".join(statuses),
        gt_label=",".join(gt_labels),
        annotation_label=",".join(annotation_labels),
        model_label=",".join(model_labels),
        missing_evidence=list(evidence_keys),
        tag_filters=legacy_tags,
        scene_tags=scene_tags,
        trigger_tags=trigger_tags,
        egress_tags=egress_tags,
        search=normalized_search,
        search_aliases=search_aliases,
    )
    tag_catalog_for_analysis = {
        str(item["key"]): {
            "label": str(item["label"]),
            "description": str(item.get("hint") or item.get("description") or ""),
            "section": str(item.get("section") or ""),
            "group": str(item.get("group") or ""),
        }
        for item in tag_catalog
    }
    result = build_review_reason_analysis(
        rows,
        theme="",
        evidence_catalog=evidence_catalog,
        tag_catalog=tag_catalog_for_analysis,
        has_model_run=bool(model_run_id),
        include_reason_themes=False,
        page=page,
        page_size=max(len(rows), 1) if unbounded else page_size,
        page_size_limit=None if unbounded else 200,
    )
    for item in result["items"]:
        issue_id = _as_text(item.get("issue_id"))
        review_params = [f"issue={quote(issue_id, safe='')}"]
        if model_run_id:
            review_params.append(f"run={quote(model_run_id, safe='')}")
        if comparison_status == "mismatch" and model_run_id:
            review_params.append("failure=1")
        item["voyager_issue_url"] = _voyager_issue_url(issue_id)
        item["review_url"] = _public_path(f"/review?{'&'.join(review_params)}")
    result["scope"] = {
        "baseline_scope": settings.baseline_scope,
        "model_run": (
            {
                "id": selected_run["id"],
                "name": selected_run["name"],
                "kind": selected_run["kind"],
                "prediction_count": selected_run["baseline_prediction_count"],
                "failure_count": selected_run["failure_count"],
            }
            if selected_run
            else None
        ),
        "comparison_status": comparison_status,
        "failure_only": comparison_status == "mismatch",
        "review_binding": "latest_annotation_per_issue",
        "review_is_run_bound": False,
    }
    result["filters"] = {
        "model_run_id": model_run_id,
        "comparison_status": comparison_status,
        "failure_only": comparison_status == "mismatch",
        "annotation_author": list(authors),
        "review_status": list(statuses),
        "gt_label": list(gt_labels),
        "annotation_label": list(annotation_labels),
        "model_label": list(model_labels),
        "missing_evidence": list(evidence_keys),
        "theme": theme,
        "tag": list(legacy_tags),
        "scene_tag": list(scene_tags),
        "trigger_tag": list(trigger_tags),
        "egress_tag": list(egress_tags),
        "search": normalized_search,
    }
    return result


@app.get("/api/review-reason-analysis")
async def review_reason_analysis(
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    annotation_author: str = "",
    review_status: str = "",
    gt_label: str = "",
    annotation_label: str = "",
    model_label: str = "",
    missing_evidence: str = "",
    theme: str = "",
    tag: str = "",
    scene_tag: str = "",
    trigger_tag: str = "",
    egress_tag: str = "",
    search: str = "",
    page: int = 1,
    page_size: int = 20,
) -> dict[str, Any]:
    return _review_reason_analysis_payload(
        model_run_id=model_run_id,
        comparison=comparison,
        failure_only=failure_only,
        annotation_author=annotation_author,
        review_status=review_status,
        gt_label=gt_label,
        annotation_label=annotation_label,
        model_label=model_label,
        missing_evidence=missing_evidence,
        theme=theme,
        tag=tag,
        scene_tag=scene_tag,
        trigger_tag=trigger_tag,
        egress_tag=egress_tag,
        search=search,
        page=page,
        page_size=page_size,
    )


REVIEW_ANALYSIS_EXPORT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("issue_id", "Issue ID"),
    ("scene", "场景"),
    ("gt_label", "GT"),
    ("model_label", "模型结论"),
    ("comparison_status", "模型判断结果"),
    ("model_reason", "模型说明"),
    ("model_confidence", "模型置信度"),
    ("review_label", "人工结论"),
    ("review_status", "Review 状态"),
    ("review_reason", "人工 Review 原因"),
    ("tags", "场景 Tags"),
    ("missing_evidence", "缺失信息"),
    ("reviewer", "复核人"),
    ("reviewed_at", "Review 时间"),
    ("review_url", "Workbench 链接"),
    ("voyager_issue_url", "Voyager Issue 链接"),
)


def _spreadsheet_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str) and value.lstrip(" \t\r\n").startswith(
        ("=", "+", "-", "@")
    ):
        return f"'{value}"
    return value


def _review_analysis_export_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    tag_labels = {str(item["key"]): str(item["label"]) for item in _review_tag_catalog()}
    evidence_labels = {
        str(item["key"]): str(item["label"])
        for item in _missing_evidence_catalog()
    }
    exported: list[dict[str, Any]] = []
    for item in result.get("items", []):
        annotation = item.get("annotation") or {}
        prediction = item.get("prediction") or {}
        exported.append(
            {
                "issue_id": _as_text(item.get("issue_id")),
                "scene": _as_text(item.get("title") or item.get("scenario")),
                "gt_label": _as_text(item.get("gt_label")),
                "model_label": _as_text(prediction.get("label")),
                "comparison_status": _as_text(item.get("comparison_status")).upper(),
                "model_reason": _as_text(prediction.get("reason")),
                "model_confidence": prediction.get("confidence"),
                "review_label": _as_text(annotation.get("label")),
                "review_status": _as_text(annotation.get("review_status")),
                "review_reason": _as_text(annotation.get("note")),
                "tags": "、".join(
                    tag_labels.get(_as_text(key), _as_text(key))
                    for key in annotation.get("tags") or []
                ),
                "missing_evidence": "、".join(
                    evidence_labels.get(_as_text(key), _as_text(key))
                    for key in annotation.get("missing_evidence") or []
                ),
                "reviewer": _as_text(annotation.get("author")),
                "reviewed_at": _as_text(annotation.get("created_at")),
                "review_url": _as_text(item.get("review_url")),
                "voyager_issue_url": _as_text(item.get("voyager_issue_url")),
            }
        )
    return exported


def _review_analysis_export_response(
    result: dict[str, Any], export_format: str
) -> Response:
    rows = _review_analysis_export_rows(result)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"review-analysis-{timestamp}.{export_format}"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    column_keys = [key for key, _ in REVIEW_ANALYSIS_EXPORT_COLUMNS]
    column_labels = [label for _, label in REVIEW_ANALYSIS_EXPORT_COLUMNS]
    if export_format == "csv":
        stream = io.StringIO()
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(column_labels)
        for row in rows:
            writer.writerow([_spreadsheet_safe(row.get(key)) for key in column_keys])
        return Response(
            content=("\ufeff" + stream.getvalue()).encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers=headers,
        )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Review 分析"
    worksheet.append(column_labels)
    for row in rows:
        worksheet.append([_spreadsheet_safe(row.get(key)) for key in column_keys])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, (_, label) in enumerate(REVIEW_ANALYSIS_EXPORT_COLUMNS, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = min(
            42, max(12, len(label) * 2 + 4)
        )
    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/review-reason-analysis/export")
async def export_review_reason_analysis(
    format: str = "csv",
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    annotation_author: str = "",
    review_status: str = "",
    gt_label: str = "",
    annotation_label: str = "",
    model_label: str = "",
    missing_evidence: str = "",
    theme: str = "",
    tag: str = "",
    scene_tag: str = "",
    trigger_tag: str = "",
    egress_tag: str = "",
    search: str = "",
) -> Response:
    export_format = _as_text(format).strip().lower()
    if export_format not in {"csv", "xlsx"}:
        raise _detail(400, "format 仅支持 csv 或 xlsx。")
    result = _review_reason_analysis_payload(
        model_run_id=model_run_id,
        comparison=comparison,
        failure_only=failure_only,
        annotation_author=annotation_author,
        review_status=review_status,
        gt_label=gt_label,
        annotation_label=annotation_label,
        model_label=model_label,
        missing_evidence=missing_evidence,
        theme=theme,
        tag=tag,
        scene_tag=scene_tag,
        trigger_tag=trigger_tag,
        egress_tag=egress_tag,
        search=search,
        unbounded=True,
    )
    return _review_analysis_export_response(result, export_format)


@app.post("/api/trail-model-sync")
async def trail_model_sync(request: Request) -> dict[str, Any]:
    try:
        body = await request.json()
    except (TypeError, ValueError):
        body = {}
    mode = _as_text(body.get("mode") or "preview")
    if mode not in {"preview", "create"}:
        raise _detail(400, "mode 仅支持 preview 或 create。")
    actor, actor_source, actor_verified = _action_actor(
        request, body.get("requested_by")
    )
    return await asyncio.to_thread(
        sync_trail_model_fields,
        create_run=mode == "create",
        requested_by=actor,
        identity_source=actor_source,
        identity_verified=actor_verified,
        trigger="manual",
    )


@app.get("/api/import-contract")
async def import_contract() -> dict[str, Any]:
    return {
        "formats": [".json", ".csv", ".xlsx", ".xlsm"],
        "issues": {
            "enabled_in_ui": False,
            "compatibility_only": True,
            "required": ["issue_id"],
            "optional": [
                "trip_id",
                "gt_label / ra_merge_result / 期望输出",
                "title",
                "scenario",
                "summary",
                "trail_url",
            ],
        },
        "model_results": {
            "required": ["issue_id", "model_label 或 ra_stuck_auto_result"],
            "optional": [
                "trip_id",
                "model_reason / reason",
                "model_confidence / confidence",
                "model_extra",
                "ra_stuck_auto_result_info",
            ],
            "native_json_envelope": {"experiment": {}, "results": []},
        },
        "autotriage_snapshot": {
            "input": "数字 Batch ID 或 records 链接",
            "source": "服务器固定只读 AutoTriage API",
            "kind": "autotriage_snapshot",
            "make_default": False,
        },
        "notes": [
            "每次导入会创建不可变 model run，并按 SHA-256 去重。",
            "页面不提供 Issue / GT 上传；旧 issues 接口仅为兼容客户端保留。Trail 真值不因模型导入而覆盖；仅显式 issues API 调用且 replace_gt=true 才会覆盖已有 GT。",
        ],
    }


async def _read_upload(upload: UploadFile) -> tuple[bytes, str, str]:
    filename = _safe_filename(upload.filename or "upload")
    content = await upload.read()
    if not content:
        raise _detail(400, "上传文件为空。")
    if len(content) > MAX_UPLOAD_BYTES:
        raise _detail(413, "上传文件超过 64 MB 限制。")
    source_hash = hashlib.sha256(content).hexdigest()
    return content, filename, source_hash


@app.post("/api/import/issues")
async def import_issues(
    file: UploadFile = File(...),
    source: str = Form("manual_upload"),
    replace_gt: str = Form("false"),
) -> dict[str, Any]:
    content, filename, _ = await _read_upload(file)
    try:
        source_rows, metadata = parse_source_bytes(filename, content)
    except (UnicodeDecodeError, ValueError, json.JSONDecodeError) as exc:
        raise _detail(400, f"解析文件失败: {exc}")
    rows = [normalized for row in source_rows if (normalized := normalize_issue_row(row, source))]
    if not rows:
        raise _detail(400, "未找到有效 issue_id；请检查表头或导入契约。")
    outcome = database.upsert_issues(
        rows,
        source=source,
        replace_gt=replace_gt.lower() in {"1", "true", "yes", "on"},
    )
    return {
        "filename": filename,
        "parsed_rows": len(source_rows),
        "accepted_rows": len(rows),
        "metadata": metadata,
        **outcome,
    }


@app.post("/api/import/model-results")
async def import_model_results(
    request: Request,
    file: UploadFile = File(...),
    run_name: str = Form(""),
    created_by: str = Form(""),
) -> dict[str, Any]:
    content, filename, source_hash = await _read_upload(file)
    try:
        source_rows, metadata = parse_source_bytes(filename, content)
    except (UnicodeDecodeError, ValueError, json.JSONDecodeError) as exc:
        raise _detail(400, f"解析文件失败: {exc}")
    rows = [normalized for row in source_rows if (normalized := normalize_model_row(row))]
    if not rows:
        raise _detail(
            400,
            "未找到有效结果；需要 issue_id 和 model_label（或 ra_stuck_auto_result）。",
        )
    try:
        source_artifact = await asyncio.to_thread(
            _store_model_source,
            content,
            filename,
            source_hash,
        )
    except OSError as exc:
        raise _detail(507, f"无法归档模型结果原文件：{exc}")
    metadata = dict(metadata)
    metadata["source_artifact"] = source_artifact
    experiment = metadata.get("experiment") if isinstance(metadata.get("experiment"), dict) else {}
    default_name = _as_text(experiment.get("model_name")) or Path(filename).stem
    actor, actor_source, actor_verified = _action_actor(request, created_by)
    run, duplicate = database.import_model_run(
        name=_as_text(run_name) or default_name,
        source_name=filename,
        source_sha256=source_hash,
        metadata=metadata,
        rows=rows,
        created_by=actor,
        created_by_source=actor_source,
        created_by_verified=actor_verified,
    )
    return {
        "run": run,
        "duplicate": duplicate,
        "filename": filename,
        "parsed_rows": len(source_rows),
        "accepted_rows": len(rows),
        "rejected_rows": len(source_rows) - len(rows),
    }


@app.get("/api/import/autotriage/{batch_id}")
async def preview_autotriage_import(batch_id: str) -> dict[str, Any]:
    try:
        normalized_batch_id = normalise_batch_id(batch_id)
    except AutoTriageSourceError as exc:
        raise _detail(400, str(exc))
    try:
        snapshot = await asyncio.to_thread(
            _fetch_autotriage_snapshot,
            normalized_batch_id,
        )
    except AutoTriageSourceError as exc:
        raise _detail(502, str(exc))
    return {
        "batch_id": snapshot["batch_id"],
        "batch": snapshot["batch"],
        "coverage": snapshot["coverage"],
        "record_url": _autotriage_record_url(snapshot["batch_id"]),
        "default_changed": False,
        "read_only": True,
    }


@app.post("/api/import/autotriage")
async def import_autotriage_results(
    request: Request,
) -> dict[str, Any]:
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "AutoTriage 导入请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "AutoTriage 导入请求必须是 JSON 对象。")
    unknown = sorted(set(body) - {"batch_id", "run_name", "created_by"})
    if unknown:
        raise _detail(400, f"AutoTriage 导入包含未知字段：{', '.join(unknown)}。")
    try:
        normalized_batch_id = normalise_batch_id(body.get("batch_id"))
    except AutoTriageSourceError as exc:
        raise _detail(400, str(exc))
    try:
        snapshot = await asyncio.to_thread(
            _fetch_autotriage_snapshot,
            normalized_batch_id,
        )
    except AutoTriageSourceError as exc:
        raise _detail(502, str(exc))
    run_name = _as_text(body.get("run_name"))
    if len(run_name) > 120:
        raise _detail(400, "Run 名称不能超过 120 个字符。")
    batch = snapshot["batch"]
    actor, actor_source, actor_verified = _action_actor(
        request,
        body.get("created_by"),
    )
    metadata = {
        "schema_version": "autotriage-snapshot-v1",
        "origin": "autotriage_readonly_snapshot",
        "source_ref": f"autotriage_batch:{snapshot['batch_id']}",
        "record_url": _autotriage_record_url(snapshot["batch_id"]),
        "platform_batch": batch,
        "coverage": snapshot["coverage"],
        "external_username": _as_text(batch.get("username")),
        "model_name": _as_text(batch.get("model_name")),
        "prompt_version": _as_text(batch.get("prompt_version")),
        "prompt_sha256": _as_text(batch.get("prompt_sha256")),
        "default_changed": False,
    }
    run, duplicate = database.import_model_run(
        name=run_name
        or _as_text(batch.get("batch_name"))
        or f"AutoTriage Batch {snapshot['batch_id']}",
        source_name=f"AutoTriage Batch {snapshot['batch_id']}",
        source_sha256=snapshot["source_sha256"],
        metadata=metadata,
        rows=snapshot["rows"],
        kind="autotriage_snapshot",
        make_default=False,
        created_by=actor,
        created_by_source=actor_source,
        created_by_verified=actor_verified,
    )
    return {
        "run": run,
        "duplicate": duplicate,
        "batch_id": snapshot["batch_id"],
        "coverage": snapshot["coverage"],
        "record_url": _autotriage_record_url(snapshot["batch_id"]),
        "default_changed": False,
        "read_only": True,
    }


@app.get("/api/prediction-batches/models")
async def batch_prediction_models(
    refresh: bool = False,
    provider_id: str = "kylin",
) -> dict[str, Any]:
    try:
        return await asyncio.to_thread(
            model_catalog.list_models,
            refresh=refresh,
            allow_stale=True,
            provider_id=provider_id,
        )
    except ModelCatalogError as exc:
        raise _detail(exc.status_code, exc.public_message)


@app.get("/api/prediction-batches/providers")
async def batch_prediction_providers() -> dict[str, Any]:
    return model_catalog.provider_catalog()


@app.get("/api/prediction-batches/prompts")
async def batch_prediction_prompts() -> dict[str, Any]:
    try:
        return await asyncio.to_thread(
            prompt_catalog.list_prompts,
            include_template=True,
        )
    except PromptCatalogError as exc:
        raise _detail(503, str(exc))


@app.get("/api/prediction-batches/config")
async def batch_prediction_config() -> dict[str, Any]:
    latest = database.list_batch_prediction_jobs(page_size=1).get("items", [])
    latest_job = latest[0] if latest else {}
    return {
        "enabled": settings.batch_prediction_enabled,
        "autotriage_push_enabled": settings.autotriage_push_enabled,
        "max_issues": settings.batch_max_issues,
        "model": {
            "source": "服务器模型网关 + 用户选择的 Prompt/Camera 输入快照",
            "name": _as_text(latest_job.get("model_name"))
            or settings.ra_model_default_id,
            "prompt_version": _as_text(latest_job.get("prompt_version"))
            or "任务创建时固化服务器 Prompt",
        },
        "model_gateway": model_catalog.status(),
        "providers": model_catalog.provider_catalog(),
        "prompt_policy": {
            "source": "cloud_server ra_auto_triage/vlm/prompts/versions",
            "editable": True,
            "immutable_per_job": True,
            "max_bytes": MAX_PROMPT_BYTES,
        },
        "input_policy": {
            "issue_source": "Voyager Issue + dashboard isolated bag cache",
            "profiles": list(INPUT_PRESETS),
            "editable_fields": [
                "frame_offsets_ms",
                "use_ra_event",
                "use_ra_options",
                "use_bev_animation",
            ],
            "frame_count_max": MAX_FRAME_COUNT,
            "frame_offset_min_ms": MIN_FRAME_OFFSET_MS,
            "frame_offset_max_ms": MAX_FRAME_OFFSET_MS,
            "ares_animation_input": True,
            "ares_animation_policy": "server_default_stuck_triage_auto_opt_api",
            "ares_capture_input": False,
            "browser_model_credentials": False,
            "bag_cache_read_only": False,
            "bag_cache_scope": "dashboard_isolated",
            "trail_write_enabled": False,
        },
        "publish_policy": {
            "explicit_confirmation": True,
            "destination": settings.auto_triage_record_base_url,
            "writer": "cloud_server 固定服务身份",
            "requester_is_writer": False,
        },
    }


@app.post("/api/prediction-batches", status_code=202)
async def create_batch_prediction(request: Request) -> dict[str, Any]:
    if not settings.batch_prediction_enabled:
        raise _detail(503, "当前部署未启用 Batch 预测。")
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "Batch 请求必须是 JSON。")
    if not isinstance(body, dict):
        raise _detail(400, "Batch 请求必须是 JSON 对象。")
    forbidden_model_fields = sorted(
        {
            "api_key",
            "base_url",
            "provider",
            "model_name",
            "model_override",
            "use_ares_capture",
            "use_bev_animation",
            "bev_mode",
            "bag_path",
            "camera_topic",
            "camera_topics",
            "experiment",
            "experiment_id",
            "experiment_revision_id",
        }.intersection(body)
    )
    if forbidden_model_fields:
        raise _detail(
            400,
            "禁止提交模型地址、凭证、provider、路径、实验 ID 或 Ares/BEV 配置。",
        )
    allowed_fields = {
        "name",
        "issue_ids",
        "requested_by",
        "provider_id",
        "model_id",
        "prompt_id",
        "prompt_template",
        "input_config",
        "allow_experimental_model",
    }
    unknown_fields = sorted(set(body) - allowed_fields)
    if unknown_fields:
        raise _detail(
            400,
            f"Batch 请求包含未知字段：{', '.join(unknown_fields)}。",
        )
    raw_issue_ids = body.get("issue_ids")
    if not isinstance(raw_issue_ids, list):
        raise _detail(400, "issue_ids 必须是数组。")
    issue_ids: list[str] = []
    seen: set[str] = set()
    for value in raw_issue_ids:
        issue_id = _as_text(value)
        if not ISSUE_ID_RE.fullmatch(issue_id):
            raise _detail(400, f"Issue ID 格式非法: {issue_id or '<空>'}")
        if issue_id not in seen:
            seen.add(issue_id)
            issue_ids.append(issue_id)
    if not issue_ids:
        raise _detail(400, "至少需要一个 Issue ID。")
    if len(issue_ids) > settings.batch_max_issues:
        raise _detail(
            400,
            f"单批最多 {settings.batch_max_issues} 个 Issue；当前 {len(issue_ids)} 个。",
        )
    provider_id = _as_text(body.get("provider_id") or "kylin").lower()
    provider_catalog = model_catalog.provider_catalog()
    provider = next(
        (
            item
            for item in provider_catalog.get("providers", [])
            if str(item.get("id") or "") == provider_id
        ),
        None,
    )
    if not provider or not provider.get("enabled"):
        raise _detail(400, "所选 Provider 未在 cloud_server 登记可用的服务端凭证。")
    missing = [
        issue_id
        for issue_id in issue_ids
        if database.get_case(issue_id) is None
    ]
    if missing:
        preview = "、".join(missing[:8])
        suffix = "…" if len(missing) > 8 else ""
        raise _detail(
            404,
            f"以下 Issue 不在当前 Workbench 数据集中: {preview}{suffix}",
        )
    name = _as_text(body.get("name"))
    if len(name) > 120:
        raise _detail(400, "Batch 名称不能超过 120 个字符。")
    try:
        model_selection = await asyncio.to_thread(
            model_catalog.resolve,
            _as_text(body.get("model_id")),
            provider_id,
        )
    except ModelCatalogError as exc:
        raise _detail(exc.status_code, exc.public_message)
    validation_status = _as_text(
        model_selection.get("validation_status")
    ) or "validated"
    if (
        validation_status != "validated"
        and body.get("allow_experimental_model") is not True
    ):
        raise _detail(
            400,
            "该 Qwen3 模型在线但尚未完成 RA 基线验证；"
            "如需试跑，请显式确认 allow_experimental_model=true。",
        )
    try:
        prompt_selection = await asyncio.to_thread(
            prompt_catalog.resolve,
            body.get("prompt_id"),
            body.get("prompt_template"),
        )
        input_config = normalise_input_config(body.get("input_config"))
    except PromptCatalogError as exc:
        raise _detail(400, str(exc))
    actor, actor_source, actor_verified = _action_actor(
        request, body.get("requested_by")
    )
    if not name:
        actor_slug = re.sub(r"[^A-Za-z0-9._-]+", "-", actor).strip("-")[:48] or "triage"
        name = f"{actor_slug}_i_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    job = database.create_batch_prediction_job(
        name=name,
        issue_ids=issue_ids,
        requested_by=actor,
        requested_by_source=actor_source,
        requested_by_verified=actor_verified,
        provider_id=provider_id,
        requested_model_id=model_selection["requested_model_id"],
        resolved_model_id=model_selection["resolved_model_id"],
        model_source="ra_model_gateway",
        catalog_sha256=model_selection["catalog_sha256"],
        model_validation_status=validation_status,
        prompt_version=prompt_selection["prompt_version"],
        prompt_template=prompt_selection["prompt_template"],
        prompt_template_sha256=prompt_selection[
            "prompt_template_sha256"
        ],
        prompt_mode=prompt_selection["prompt_mode"],
        input_profile=input_config["profile_id"],
        input_config=input_config,
    )
    if not batch_prediction_runner.launch_prediction(job):
        error = "Batch worker 正在停止，任务无法入队；请稍后重试。"
        database.update_batch_prediction_items(
            job["id"],
            [
                {"issue_id": issue_id, "success": False, "error": error}
                for issue_id in issue_ids
            ],
        )
        database.update_batch_prediction_job(
            job["id"],
            status="failed",
            completed_count=len(issue_ids),
            failed_count=len(issue_ids),
            error_text=error,
        )
        raise _detail(409, error)
    return {
        "job": _public_batch_job(
            database.get_batch_prediction_job(job["id"]) or job
        ),
        "safety": {
            "server_default_model": False,
            "server_model_gateway": True,
            "provider_id": provider_id,
            "requested_model_id": model_selection["requested_model_id"],
            "resolved_model_id": model_selection["resolved_model_id"],
            "model_validation_status": validation_status,
            "prompt_version": prompt_selection["prompt_version"],
            "prompt_sha256": prompt_selection["prompt_template_sha256"],
            "prompt_mode": prompt_selection["prompt_mode"],
            "input_profile": input_config["profile_id"],
            "browser_model_credentials": False,
            "ares_bev_input": bool(input_config["use_bev_animation"]),
            "bag_cache_read_only": False,
            "bag_cache_scope": "dashboard_isolated",
            "trail_write_enabled": False,
            "autotriage_publish_automatic": False,
        },
        "poll_url": _public_path(f"/api/prediction-batches/{job['id']}"),
    }


@app.get("/api/prediction-batches")
async def list_batch_predictions(
    requested_by: str = "",
    status: str = "",
    model_id: str = "",
    prompt_version: str = "",
    prompt_mode: str = "",
    prompt_sha256: str = "",
    input_profile: str = "",
    page_size: int = 100,
) -> dict[str, Any]:
    result = await asyncio.to_thread(
        database.list_batch_prediction_jobs,
        requested_by=requested_by,
        status=status,
        model_id=model_id,
        prompt_version=prompt_version,
        prompt_mode=prompt_mode,
        prompt_sha256=prompt_sha256,
        input_profile=input_profile,
        page_size=page_size,
    )
    result["items"] = [
        _public_batch_job(job) for job in result.get("items", [])
    ]
    return result


@app.get("/api/prediction-batches/{job_id}")
async def get_batch_prediction(job_id: str) -> dict[str, Any]:
    job = await asyncio.to_thread(database.get_batch_prediction_job, job_id)
    if job is None:
        raise _detail(404, "Batch 任务不存在。")
    return {"job": _public_batch_job(job, include_prompt=True)}


@app.post(
    "/api/prediction-batches/{job_id}/publish-autotriage",
    status_code=202,
)
async def publish_batch_prediction(job_id: str, request: Request) -> dict[str, Any]:
    if request.headers.get("x-ra-triage-request") != "publish-v1":
        raise _detail(403, "缺少 AutoTriage 推送请求标记。")
    if not settings.autotriage_push_enabled:
        raise _detail(503, "当前部署未启用 AutoTriage 推送。")
    publisher = request_identity(request, settings)
    if not publisher.verified or not publisher.username:
        raise _detail(
            403,
            "AutoTriage 是生产写操作，仅允许可信 SSO 会话；"
            "直接 IP / 本机 LCA 用户不能触发。",
        )
    try:
        body = await request.json()
    except (TypeError, ValueError):
        raise _detail(400, "推送请求必须是 JSON。")
    if not isinstance(body, dict) or body.get("confirm") not in {
        True,
        "publish-autotriage",
    }:
        raise _detail(400, "必须显式确认创建生产 AutoTriage Batch。")
    job = database.get_batch_prediction_job(job_id)
    if job is None:
        raise _detail(404, "Batch 任务不存在。")
    if job.get("autotriage_batch_id"):
        return {"job": _public_batch_job(job), "idempotent": True}
    if job.get("publish_status") == "running":
        raise _detail(409, "该 Batch 正在推送 AutoTriage。")
    if not MODEL_ID_RE.fullmatch(_as_text(job.get("resolved_model_id"))):
        raise _detail(
            409,
            "该历史 Batch 缺少已解析模型信息，不能安全推送；请重新发起预测。",
        )
    if (
        not _as_text(job.get("prompt_template"))
        or not re.fullmatch(
            r"[0-9a-f]{64}",
            _as_text(job.get("prompt_template_sha256")).lower(),
        )
        or not _as_text(job.get("input_profile"))
        or not isinstance(job.get("input_config"), dict)
    ):
        raise _detail(
            409,
            "该历史 Batch 缺少不可变 Prompt/Input 快照，"
            "不能重建同一配置推送；请重新发起预测。",
        )
    if (
        job.get("status") not in {"succeeded", "partial"}
        or int(job.get("success_count") or 0) <= 0
        or not job.get("model_run_id")
        or not job.get("config_sha256")
    ):
        raise _detail(409, "Batch 尚无可推送的成功预测。")
    job = (
        database.update_batch_prediction_job(
            job_id,
            summary={
                **dict(job.get("summary") or {}),
                "autotriage_publish_request": {
                    "requested_by": publisher.username,
                    "identity_source": publisher.source,
                    "verified": True,
                },
            },
        )
        or job
    )
    if not batch_prediction_runner.launch_publish(job):
        raise _detail(409, "已有 Batch 预测或 AutoTriage 推送正在执行，请稍后重试。")
    return {
        "job": _public_batch_job(
            database.get_batch_prediction_job(job_id) or job
        ),
        "accepted": True,
        "destination": settings.auto_triage_record_base_url,
        "poll_url": _public_path(f"/api/prediction-batches/{job_id}"),
    }


@app.post("/api/inference/jobs")
async def create_inference_job(request: Request) -> dict[str, Any]:
    raise _detail(
        410,
        "浏览器单 Case 自定义模型推理已停用；请使用 /api/prediction-batches，"
        "由 cloud_server 的 ra_auto_triage 默认模型执行。",
    )


@app.get("/api/inference/jobs")
async def list_inference_jobs(
    requested_by: str = "",
    status: str = "",
    page_size: int = 100,
) -> dict[str, Any]:
    return database.list_inference_jobs(
        requested_by=requested_by,
        status=status,
        page_size=page_size,
    )


@app.get("/api/inference/jobs/{job_id}")
async def get_inference_job(job_id: str) -> dict[str, Any]:
    job = database.get_job(job_id)
    if job is None:
        raise _detail(404, "任务不存在。")
    return {"job": job}
    identity_header_candidates,

from __future__ import annotations

import asyncio
import csv
import io
from datetime import datetime
from typing import Any

import openpyxl
from fastapi import APIRouter, Request
from fastapi.responses import Response

from ..db import LABELS
from ..http_support import (
    _as_text,
    _detail,
    _missing_evidence_catalog,
    _review_reason_analysis_payload,
    _review_tag_catalog,
    resolve_request_baseline_ids,
    resolve_request_baseline_scopes,
)

router = APIRouter()

@router.get("/api/review-reason-analysis")
async def review_reason_analysis(
    request: Request,
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    annotation_author: str = "",
    review_status: str = "",
    gt_label: str = "",
    annotation_label: str = "",
    model_label: str = "",
    missing_evidence: str = "",
    theme: str = "",
    tag: str = "",
    scene_tag: str = "",
    trigger_tag: str = "",
    egress_tag: str = "",
    search: str = "",
    page: int = 1,
    page_size: int = 20,
    baselines: str = "",
) -> dict[str, Any]:
    scopes = resolve_request_baseline_scopes(baselines, request=request)
    payload = await asyncio.to_thread(
        _review_reason_analysis_payload,
        model_run_id=model_run_id,
        comparison=comparison,
        failure_only=failure_only,
        annotation_author=annotation_author,
        review_status=review_status,
        gt_label=gt_label,
        annotation_label=annotation_label,
        model_label=model_label,
        missing_evidence=missing_evidence,
        theme=theme,
        tag=tag,
        scene_tag=scene_tag,
        trigger_tag=trigger_tag,
        egress_tag=egress_tag,
        search=search,
        page=page,
        page_size=page_size,
        baselines=baselines,
        baseline_scopes=scopes,
    )
    payload["baselines"] = resolve_request_baseline_ids(baselines, request=request)
    payload["baseline_scopes"] = scopes
    return payload


REVIEW_ANALYSIS_EXPORT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("issue_id", "Issue ID"),
    ("scene", "场景"),
    ("gt_label", "GT"),
    ("model_label", "模型结论"),
    ("comparison_status", "模型判断结果"),
    ("model_reason", "模型说明"),
    ("model_confidence", "模型置信度"),
    ("expected_output", "期望输出"),
    ("review_status", "自动状态"),
    ("review_reason", "人工 Review 原因"),
    ("tags", "场景 Tags"),
    ("missing_evidence", "缺失信息"),
    ("reviewer", "复核人"),
    ("reviewed_at", "Review 时间"),
    ("review_url", "Workbench 链接"),
    ("voyager_issue_url", "Voyager Issue 链接"),
)


def _spreadsheet_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str) and value.lstrip(" \t\r\n").startswith(
        ("=", "+", "-", "@")
    ):
        return f"'{value}"
    return value


def _review_analysis_export_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    tag_labels = {str(item["key"]): str(item["label"]) for item in _review_tag_catalog()}
    evidence_labels = {
        str(item["key"]): str(item["label"])
        for item in _missing_evidence_catalog()
    }
    exported: list[dict[str, Any]] = []
    for item in result.get("items", []):
        annotation = item.get("annotation") or {}
        prediction = item.get("prediction") or {}
        expected_output = _as_text(
            annotation.get("expected_output") or annotation.get("label")
        )
        exported.append(
            {
                "issue_id": _as_text(item.get("issue_id")),
                "scene": _as_text(item.get("title") or item.get("scenario")),
                "gt_label": _as_text(item.get("gt_label")),
                "model_label": _as_text(prediction.get("label")),
                "comparison_status": _as_text(item.get("comparison_status")).upper(),
                "model_reason": _as_text(prediction.get("reason")),
                "model_confidence": prediction.get("confidence"),
                "expected_output": expected_output,
                "review_status": _as_text(annotation.get("review_status")),
                "review_reason": _as_text(annotation.get("note")),
                "tags": "、".join(
                    tag_labels.get(_as_text(key), _as_text(key))
                    for key in annotation.get("tags") or []
                ),
                "missing_evidence": "、".join(
                    evidence_labels.get(_as_text(key), _as_text(key))
                    for key in annotation.get("missing_evidence") or []
                ),
                "reviewer": _as_text(annotation.get("author")),
                "reviewed_at": _as_text(annotation.get("created_at")),
                "review_url": _as_text(item.get("review_url")),
                "voyager_issue_url": _as_text(item.get("voyager_issue_url")),
            }
        )
    return exported


def _trail_expected_output_rows(result: dict[str, Any]) -> list[dict[str, str]]:
    """Return only GT-changing rows accepted by 张扬's expected-output mode."""

    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in result.get("items", []):
        annotation = item.get("annotation") or {}
        issue_id = _as_text(item.get("issue_id"))
        expected_output = _as_text(
            annotation.get("expected_output") or annotation.get("label")
        )
        gt_label = _as_text(item.get("gt_label"))
        if (
            not issue_id
            or issue_id in seen
            or expected_output not in LABELS
            or expected_output == gt_label
        ):
            continue
        seen.add(issue_id)
        rows.append({"issue_id": issue_id, "期望输出": expected_output})
    return rows


def _review_analysis_export_response(
    result: dict[str, Any], export_format: str
) -> Response:
    if export_format == "trail_xlsx":
        rows = _trail_expected_output_rows(result)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "GT 更新"
        worksheet.append(["issue_id", "期望输出"])
        for row in rows:
            worksheet.append(
                [
                    _spreadsheet_safe(row["issue_id"]),
                    _spreadsheet_safe(row["期望输出"]),
                ]
            )
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 16
        output = io.BytesIO()
        workbook.save(output)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": (
                    f'attachment; filename="gt-update-{timestamp}.xlsx"'
                )
            },
        )

    rows = _review_analysis_export_rows(result)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"review-analysis-{timestamp}.{export_format}"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    column_keys = [key for key, _ in REVIEW_ANALYSIS_EXPORT_COLUMNS]
    column_labels = [label for _, label in REVIEW_ANALYSIS_EXPORT_COLUMNS]
    if export_format == "csv":
        stream = io.StringIO()
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(column_labels)
        for row in rows:
            writer.writerow([_spreadsheet_safe(row.get(key)) for key in column_keys])
        return Response(
            content=("\ufeff" + stream.getvalue()).encode("utf-8"),
            media_type="text/csv; charset=utf-8",
            headers=headers,
        )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Review 分析"
    worksheet.append(column_labels)
    for row in rows:
        worksheet.append([_spreadsheet_safe(row.get(key)) for key in column_keys])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, (_, label) in enumerate(REVIEW_ANALYSIS_EXPORT_COLUMNS, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = min(
            42, max(12, len(label) * 2 + 4)
        )
    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )



@router.get("/api/review-reason-analysis/export")
async def export_review_reason_analysis(
    request: Request,
    format: str = "csv",
    model_run_id: str = "",
    comparison: str = "",
    failure_only: bool = False,
    annotation_author: str = "",
    review_status: str = "",
    gt_label: str = "",
    annotation_label: str = "",
    model_label: str = "",
    missing_evidence: str = "",
    theme: str = "",
    tag: str = "",
    scene_tag: str = "",
    trigger_tag: str = "",
    egress_tag: str = "",
    search: str = "",
    baselines: str = "",
) -> Response:
    export_format = _as_text(format).strip().lower()
    if export_format not in {"csv", "xlsx", "trail_xlsx"}:
        raise _detail(400, "format 仅支持 csv、xlsx 或 trail_xlsx。")
    scopes = resolve_request_baseline_scopes(baselines, request=request)
    result = await asyncio.to_thread(
        _review_reason_analysis_payload,
        model_run_id=model_run_id,
        comparison=comparison,
        failure_only=failure_only,
        annotation_author=annotation_author,
        review_status=review_status,
        gt_label=gt_label,
        annotation_label=annotation_label,
        model_label=model_label,
        missing_evidence=missing_evidence,
        theme=theme,
        tag=tag,
        scene_tag=scene_tag,
        trigger_tag=trigger_tag,
        egress_tag=egress_tag,
        search=search,
        unbounded=True,
        baselines=baselines,
        baseline_scopes=scopes,
    )
    result["baselines"] = resolve_request_baseline_ids(baselines, request=request)
    return await asyncio.to_thread(
        _review_analysis_export_response, result, export_format
    )

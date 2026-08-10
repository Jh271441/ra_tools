from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from .db import LABELS


@dataclass(frozen=True)
class BaselineLoad:
    rows: list[dict[str, Any]]
    source_rows: int
    skipped_rows: int
    message: str


# Chinese / free-form labels seen in spotcheck workbooks → three-class GT.
_LABEL_ALIASES = {
    "误触发": "误触发",
    "false": "误触发",
    "false_trigger": "误触发",
    "fp": "误触发",
    "正确触发": "正确触发",
    "成功": "正确触发",
    "失败": "正确触发",
    "true": "正确触发",
    "true_trigger": "正确触发",
    "tp": "正确触发",
    "无需协助": "无需协助",
    "no_assist": "无需协助",
    "no assist": "无需协助",
    "无需": "无需协助",
}


def normalize_gt_label(raw: Any) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    if text in LABELS:
        return text
    return _LABEL_ALIASES.get(text) or _LABEL_ALIASES.get(text.lower()) or ""


def load_label_baseline(path: Path, dataset: str) -> BaselineLoad:
    """Load the stable workset and rollback GT seed without live Trail I/O.

    The persisted authoritative GT overlay is merged by the database bootstrap
    layer.  Membership still comes from this dataset-filtered workbook, so the
    0508 workset remains the same 1071 cases across refreshes and restarts.
    """

    if not path.is_file():
        return BaselineLoad([], 0, 0, f"基线文件不存在: {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "labels" not in workbook.sheetnames:
        return BaselineLoad([], 0, 0, "基线 Excel 缺少 labels sheet。")
    sheet = workbook["labels"]
    values = sheet.iter_rows(values_only=True)
    headers = next(values, None)
    if not headers:
        return BaselineLoad([], 0, 0, "基线 Excel 缺少表头。")
    index = {str(value).strip(): position for position, value in enumerate(headers) if value}
    required = {"dataset", "issue_id", "Final Label"}
    missing = sorted(required - set(index))
    if missing:
        return BaselineLoad([], 0, 0, f"基线 Excel 缺少列: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    source_rows = skipped_rows = 0
    for raw in values:
        if not raw or not any(value not in (None, "") for value in raw):
            continue
        if str(raw[index["dataset"]] or "").strip() != dataset:
            continue
        source_rows += 1
        issue_id = str(raw[index["issue_id"]] or "").strip()
        label = normalize_gt_label(raw[index["Final Label"]])
        if not issue_id or label not in LABELS:
            skipped_rows += 1
            continue
        extra = {
            key: raw[position]
            for key, position in index.items()
            if position < len(raw) and raw[position] not in (None, "")
        }
        rows.append(
            {
                "issue_id": issue_id,
                "gt_label": label,
                "gt_source": f"{path.name}:labels.dataset={dataset}",
                "scenario": str(extra.get("source_dataset_label") or "").strip(),
                "extra": {"baseline": extra},
            }
        )
    message = f"已读取 {dataset} 基线 {len(rows)} 条"
    if skipped_rows:
        message += f"，跳过 {skipped_rows} 条无效记录"
    return BaselineLoad(rows, source_rows, skipped_rows, message)


def _sheet_header_index(headers: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    return {
        str(value).strip(): position
        for position, value in enumerate(headers)
        if value not in (None, "")
    }


def _pick_column(index: dict[str, int], *candidates: str) -> str | None:
    for name in candidates:
        if name in index:
            return name
    lowered = {key.lower(): key for key in index}
    for name in candidates:
        hit = lowered.get(name.lower())
        if hit:
            return hit
    return None


def load_spotcheck_zh_baseline(path: Path) -> BaselineLoad:
    """Load a Chinese spot-check workbook (0626 抽检 style).

    Expected columns (any one of the aliases):
    - issue: 问题 ID / issue_id / Issue ID
    - label: 合并标注 / Final Label / gt_label
    - comment (optional): 最终Comment / Comment / summary
    """

    if not path.is_file():
        return BaselineLoad([], 0, 0, f"基线文件不存在: {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    values = sheet.iter_rows(values_only=True)
    headers = next(values, None)
    if not headers:
        return BaselineLoad([], 0, 0, "抽检 Excel 缺少表头。")
    index = _sheet_header_index(headers)
    issue_col = _pick_column(index, "问题 ID", "问题ID", "issue_id", "Issue ID", "IssueId")
    label_col = _pick_column(
        index, "合并标注", "Final Label", "gt_label", "标注", "最终标注"
    )
    comment_col = _pick_column(
        index, "最终Comment", "最终 Comment", "Comment", "comment", "summary", "备注"
    )
    if not issue_col or not label_col:
        return BaselineLoad(
            [],
            0,
            0,
            f"抽检 Excel 缺少列（需要问题 ID / 合并标注）: {', '.join(sorted(index))}",
        )

    rows: list[dict[str, Any]] = []
    source_rows = skipped_rows = 0
    for raw in values:
        if not raw or not any(value not in (None, "") for value in raw):
            continue
        source_rows += 1
        issue_id = str(raw[index[issue_col]] or "").strip()
        label = normalize_gt_label(raw[index[label_col]])
        if not issue_id or label not in LABELS:
            skipped_rows += 1
            continue
        summary = ""
        if comment_col is not None:
            summary = str(raw[index[comment_col]] or "").strip()
        extra = {
            key: raw[position]
            for key, position in index.items()
            if position < len(raw) and raw[position] not in (None, "")
        }
        rows.append(
            {
                "issue_id": issue_id,
                "gt_label": label,
                "gt_source": path.name,
                "summary": summary,
                "extra": {"baseline": extra, "loader": "spotcheck_zh"},
            }
        )
    message = f"已读取抽检基线 {len(rows)} 条"
    if skipped_rows:
        message += f"，跳过 {skipped_rows} 条无效记录"
    return BaselineLoad(rows, source_rows, skipped_rows, message)


def load_baseline_entry(
    *,
    loader: str,
    path: Path,
    dataset: str = "",
) -> BaselineLoad:
    if loader == "trail_label_baseline":
        if not dataset:
            return BaselineLoad([], 0, 0, "trail_label_baseline 需要 dataset。")
        return load_label_baseline(path, dataset)
    if loader == "spotcheck_zh":
        return load_spotcheck_zh_baseline(path)
    return BaselineLoad([], 0, 0, f"未知 baseline loader: {loader}")

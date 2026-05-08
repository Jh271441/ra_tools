#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

import numpy as np
from enum import Enum
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

WEB_MONITOR_BASE = (
    "https://voyager.intra.xiaojukeji.com/static/ares-studio/"
    "?ds=voy-ws-car&ds.server=&layoutId=79498cb5-24dd-4585-83d1-2b5be6f8589a"
)


class AssistModelProcessReason(Enum):
    UNDEFINED_REASON = 0
    FP_TRAFFIC_JAM = 1
    FP_RED_LIGHT = 2
    FP_STOP_FENCE = 3
    FP_STARTUP = 4
    FP_MAXSPD = 5
    FP_CROSSWALK = 6
    FN_REJECT_INQUEUE = 7
    FN_JUNCTION = 8
    FN_SOLID_LANEMARK = 9
    FN_LOW_PRED = 10
    FP_PULL_OVER = 11
    FP_YIELD_DYNAMIC_OBJECT = 12
    FP_EOL_WITH_RED_TL = 13
    FP_YIELD_ON_RIGHT_TURN = 14
    FP_QUEUING = 15
    FN_RA_CZ = 16
    FP_YIELD_ON_TURN = 17
    FN_NEAR_HARD_BOUNDARY = 18
    FN_SELECTION = 19
    FP_OCCLUSION = 20
    FN_BREAKDOWN_CAR = 21
    RESERVED = 22
    REQUEST_FROM_ROUTING = 23
    FN_PERCEPTION_FP = 24
    FN_EOL = 25
    FP_REMOTE_SPEED_LIMIT = 26
    FN_CZ = 27
    REQUEST_FROM_CREEP = 28
    FN_NO_BLOCK = 29
    FN_LANE_CHANGE_STUCK = 30
    FN_FORCING_RECALL = 31
    FN_VEHICLE_HAZARD_SIGNAL = 32
    REQUEST_FROM_TIDAL_FLOW_LANE = 33
    REQUIREMENT_OF_TRAFFIC_LIGHT = 34
    FN_ABNORMAL_TRAFFIC_LIGHT = 35
    ASSIST_STUCK_MODEL = 36
    SPECIAL_STUCK_SCENE = 37
    FP_OPEN_SPACE_PLANNING = 38
    FP_LANE_CHANGE_FORBID = 39
    FN_FINAL_FORCING_RECALL = 40


def load_single_npz(path):
    try:
        data = np.load(path, allow_pickle=True)
        if "segments" in data.files:
            segments = data["segments"].tolist()
        else:
            segments = [{k: data[k].item() if data[k].shape == () else data[k].tolist()
                         for k in data.files}]
        return segments
    except Exception as e:
        print(f"Failed to load {path}: {e}")
        return []


def load_all_segments(all_files, max_workers=16):
    all_segments = []
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(load_single_npz, path): path for path in all_files}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Loading segments: "):
            segments = future.result()
            all_segments.extend(segments)
    return all_segments


def load_npz_files_from_dirs(dirs):
    all_files = []
    for root_dir in dirs:
        for root, _, files in os.walk(root_dir):
            for f in files:
                if f.endswith(".npz"):
                    path = os.path.join(root, f)
                    all_files.append(path)
    return load_all_segments(all_files)


def ns_to_rfc3339_utc(ns):
    """Convert nanoseconds since epoch to RFC3339 string in UTC with nanoseconds."""
    seconds, nanoseconds = divmod(ns, 1_000_000_000)
    dt = datetime.fromtimestamp(seconds, tz=timezone.utc)
    # 格式化为 RFC3339，保留纳秒
    rfc3339_str = dt.strftime("%Y-%m-%dT%H:%M:%S") + f".{nanoseconds:09d}Z"
    # URL encode ':' -> %3A
    return rfc3339_str.replace(":", "%3A")


def generate_web_link(trip_id, start_ns, end_ns):
    start_ms = start_ns // 1_000_000
    end_ms = int(end_ns // 1_000_000) + 1_000
    rfc3339_time = ns_to_rfc3339_utc(end_ns)
    return f"{WEB_MONITOR_BASE}&ds.start={start_ms}&ds.end={end_ms}&ds.trip_id={trip_id}&time={rfc3339_time}"


def cluster_segments_by_reason(segments):
    clustered = defaultdict(list)
    for seg in segments:
        reasons = seg.get("trigger_reasons", [])
        if not reasons:
            clustered["MANUAL_TRIGGER"].append(seg)
        else:
            for r in reasons:
                try:
                    reason_name = AssistModelProcessReason(r).name
                except ValueError:
                    reason_name = f"UNKNOWN({r})"
                clustered[reason_name].append(seg)
    return clustered


def save_summary_xlsx(clustered, output_path="summary.xlsx"):
    if not os.path.exists(os.path.dirname(output_path)):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Trigger Segments"

    # --- 新增列 headers
    headers = ["trigger_reason", "trip_id", "start_time_ns", "end_time_ns", "trigger_timestamps_ms", "web_link"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # 按 segment 数量排序
    clustered_sorted = dict(sorted(clustered.items(), key=lambda kv: len(kv[1]), reverse=True))
    for reason, segs in clustered_sorted.items():
        for seg in segs:
            link = generate_web_link(seg["trip_id"], seg["start_time"], seg["end_time"])
            trigger_times_ms = ",".join(str(int(ts)) for ts in seg.get("trigger_timestamps", []))
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=1, value=reason)
            ws.cell(row=row_idx, column=2, value=seg["trip_id"])
            ws.cell(row=row_idx, column=3, value=seg["start_time"])
            ws.cell(row=row_idx, column=4, value=seg["end_time"])
            ws.cell(row=row_idx, column=5, value=trigger_times_ms)
            # HYPERLINK
            ws.cell(row=row_idx, column=6).hyperlink = link
            ws.cell(row=row_idx, column=6).value = "View"
            ws.cell(row=row_idx, column=6).style = "Hyperlink"

    # 自动列宽
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_path)
    print(f"Saved XLSX to {output_path}")


def summarize_trigger_reasons(segments):
    reason_counter = Counter()
    empty_segments = []
    for seg in segments:
        reasons = seg.get("trigger_reasons", [])
        if not reasons:
            empty_segments.append(seg)
        for r in reasons:
            try:
                reason_name = AssistModelProcessReason(r).name
            except ValueError:
                reason_name = f"UNKNOWN({r})"
            reason_counter[reason_name] += 1
    return reason_counter, empty_segments


if __name__ == "__main__":
    dirs_to_scan = [
        "/home/luban/ofs/user/jasperchen/2026Q1_swag_trigger_scenario_clustering_from_rt_event/results/0106",
    ]

    segments = load_npz_files_from_dirs(dirs_to_scan)
    print(f"Total segments loaded: {len(segments)}")

    # 统计 trigger 原因
    reason_counter, empty_segments = summarize_trigger_reasons(segments)
    print("\n=== Trigger Reason Distribution (by segment count) ===")
    for reason, cnt in reason_counter.most_common():
        print(f"{reason:35}: {cnt}")

    print(f"\n=== Empty trigger_reasons segments: {len(empty_segments)} ===")
    for seg in empty_segments:
        print(
            f"trip_id={seg['trip_id']}, start={seg['start_time']}, "
            f"end={seg['end_time']}, trigger_count={seg['trigger_count']}"
        )

    clustered = cluster_segments_by_reason(segments)
    save_summary_xlsx(clustered, "/home/didi/workspace/ra_tools/swag/results/summary.xlsx")
